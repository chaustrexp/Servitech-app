from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from turnos.decorators import rol_requerido
from django.contrib import messages
from django.db.models import Count
from datetime import date, timedelta
import calendar
from django.http import JsonResponse
from django.views.decorators.http import require_POST

from ..models import Cita, Usuario, Repuesto, Servicio, Especialidad, EstadoCita
from ..forms import EditarPerfilForm


# ─────────────────────────────────────────────
#  DASHBOARD ADMINISTRADOR
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_dashboard(request):
    """Dashboard principal del administrador con métricas reales."""
    hoy = date.today()

    # ── KPIs ──
    total_citas     = Cita.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month).count()
    total_tecnicos  = Usuario.objects.filter(rol=Usuario.Rol.TECNICO, activo=True).count()
    total_usuarios  = Usuario.objects.filter(activo=True).count()
    total_servicios = Servicio.objects.filter(activo=True).count()

    # Citas de hoy
    citas_hoy = Cita.objects.filter(fecha=hoy).count()

    # ── Actividad reciente (últimas 6 citas) ──
    actividad_reciente = (
        Cita.objects
        .select_related('cliente', 'servicio', 'estado', 'tecnico')
        .order_by('-fecha_creacion')[:6]
    )

    # ── Repuestos con stock bajo (menos de 3 unidades) ──
    stock_bajo = Repuesto.objects.filter(activo=True, stock__lt=3).order_by('stock')[:3]

    # ── Tendencias por día de la semana (semana actual) ──
    inicio_semana = hoy - timedelta(days=hoy.weekday())  # lunes
    tendencias = []
    dias_label = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    for i in range(7):
        dia = inicio_semana + timedelta(days=i)
        total_dia = Cita.objects.filter(fecha=dia).count()
        finalizadas_dia = Cita.objects.filter(fecha=dia, estado__nombre__iexact='Finalizada').count()
        tendencias.append({
            'label': dias_label[i],
            'total': total_dia,
            'finalizadas': finalizadas_dia,
            'es_hoy': dia == hoy,
        })

    # Máximo para calcular porcentajes de las barras
    max_tendencia = max((t['total'] for t in tendencias), default=1) or 1

    context = {
        'total_citas':        total_citas,
        'total_tecnicos':     total_tecnicos,
        'total_usuarios':     total_usuarios,
        'total_servicios':    total_servicios,
        'citas_hoy':          citas_hoy,
        'actividad_reciente': actividad_reciente,
        'stock_bajo':         stock_bajo,
        'tendencias':         tendencias,
        'max_tendencia':      max_tendencia,
    }
    return render(request, 'turnos/administracion/admin_dashboard.html', context)

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def dashboard_tecnico(request):
    """Dashboard principal del técnico."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    from turnos.services.cancelacion_automatica import cancelar_citas_vencidas
    cancelar_citas_vencidas()

    from datetime import date
    hoy = date.today()

    try:
        especialidad = request.user.perfil_tecnico.especialidad
    except Exception:
        especialidad = 'tecnico_general'

    dispositivos_permitidos = []
    if especialidad == 'tecnico_celular':
        dispositivos_permitidos = ['CELULAR']
    elif especialidad == 'tecnico_pc':
        dispositivos_permitidos = ['PC']
    elif especialidad == 'tecnico_laptop':
        dispositivos_permitidos = ['LAPTOP']
    else: # tecnico_general
        filtro_dispositivo = request.GET.get('filtro_dispositivo')
        if filtro_dispositivo in ['CELULAR', 'PC', 'LAPTOP']:
            dispositivos_permitidos = [filtro_dispositivo]
        else:
            dispositivos_permitidos = ['CELULAR', 'PC', 'LAPTOP']

    base_qs = Cita.objects.filter(servicio__tipo_dispositivo__in=dispositivos_permitidos)

    # Citas programadas para hoy asignadas al técnico
    citas_hoy_count = base_qs.filter(fecha=hoy, tecnico=request.user).count()

    # Citas en proceso asignadas al técnico (En diagnóstico o En reparación)
    en_proceso_count = base_qs.filter(
        tecnico=request.user,
        estado__nombre__in=['EN_REPARACION']
    ).count()

    # Citas finalizadas asignadas al técnico
    finalizadas_count = base_qs.filter(
        tecnico=request.user,
        estado__nombre__iexact='FINALIZADA'
    ).count()

    # Citas con retraso asignadas al técnico
    retrasos_count = base_qs.filter(
        tecnico=request.user,
        estado__nombre__iexact='RETRASADA'
    ).count()

    # Citas disponibles: asignadas al técnico en estado pendiente/confirmada/reagendada
    ESTADOS_DISPONIBLES = [
        'PENDIENTE',
        'CONFIRMADA',
        'REAGENDADA',
    ]
    citas_disponibles = base_qs.filter(
        estado__nombre__in=ESTADOS_DISPONIBLES,
        tecnico=request.user
    ).select_related('cliente', 'servicio', 'estado').order_by('fecha', 'hora_inicio')

    # ────────────────────────────────────────────────────────
    # Rendimiento Semanal (Lunes a Domingo de la semana actual)
    # ────────────────────────────────────────────────────────
    from datetime import timedelta
    from django.utils import timezone
    hoy = timezone.now().date()
    inicio_semana = hoy - timedelta(days=hoy.weekday()) # Lunes
    
    NOMBRES_DIA = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    dias_semana = []
    max_citas_dia = 1
    total_semana = 0
    
    for i in range(7):
        dia = inicio_semana + timedelta(days=i)
        count = Cita.objects.filter(
            tecnico=request.user,
            fecha=dia
        ).exclude(estado__nombre__iexact='CANCELADA').count()
        dias_semana.append({
            'nombre': NOMBRES_DIA[i],
            'count': count,
            'fecha': dia
        })
        total_semana += count
        if count > max_citas_dia:
            max_citas_dia = count

    # Calcular porcentajes para la altura de las barras
    for d in dias_semana:
        d['porcentaje'] = int((d['count'] / max_citas_dia) * 100) if max_citas_dia > 0 else 0
        d['es_hoy'] = (d['fecha'] == hoy)

    # ────────────────────────────────────────────────────────
    # Tarea Actual (Cita en curso)
    # ────────────────────────────────────────────────────────
    tarea_actual = base_qs.filter(
            tecnico=request.user,
            estado__nombre__in=['EN_REPARACION']
        ).order_by('hora_inicio').first()

    progreso_tarea = 0
    pasos_tarea = []
    
    if tarea_actual:
        est = tarea_actual.estado.nombre.upper()
        if 'DIAGN' in est:
            progreso_tarea = 35
            pasos_tarea = [
                {'nombre': 'Revisión inicial', 'completado': True},
                {'nombre': 'Diagnóstico de hardware', 'completado': False},
                {'nombre': 'Presupuesto pendiente', 'completado': False},
            ]
        elif 'REPAR' in est or 'PROCESO' in est:
            progreso_tarea = 70
            pasos_tarea = [
                {'nombre': 'Diagnóstico completado', 'completado': True},
                {'nombre': 'Reparación en curso', 'completado': True},
                {'nombre': 'Pruebas finales pendientes', 'completado': False},
            ]
        else:
            progreso_tarea = 10
            pasos_tarea = [{'nombre': 'Esperando revisión', 'completado': False}]

    context = {
        'citas_hoy_count': citas_hoy_count,
        'en_proceso_count': en_proceso_count,
        'finalizadas_count': finalizadas_count,
        'retrasos_count': retrasos_count,
        'citas_disponibles': citas_disponibles,
        'total_disponibles': citas_disponibles.count(),
        'hoy': hoy,
        'dias_semana': dias_semana,
        'total_semana': total_semana,
        'tarea_actual': tarea_actual,
        'progreso_tarea': progreso_tarea,
        'pasos_tarea': pasos_tarea,
    }
    return render(request, 'turnos/tecnico/tecnico_inicio.html', context)

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_agenda(request):
    """Agenda del técnico."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    from turnos.services.cancelacion_automatica import cancelar_citas_vencidas
    cancelar_citas_vencidas()

    from datetime import date, timedelta
    hoy = date.today()

    try:
        semana_offset = int(request.GET.get('semana_offset', 0))
    except (ValueError, TypeError):
        semana_offset = 0

    # Lunes de la semana seleccionada
    inicio_semana = hoy - timedelta(days=hoy.weekday()) + timedelta(weeks=semana_offset)
    fin_semana = inicio_semana + timedelta(days=6)

    try:
        especialidad = request.user.perfil_tecnico.especialidad
    except Exception:
        especialidad = 'tecnico_general'

    dispositivos_permitidos = []
    if especialidad == 'tecnico_celular':
        dispositivos_permitidos = ['CELULAR']
    elif especialidad == 'tecnico_pc':
        dispositivos_permitidos = ['PC']
    elif especialidad == 'tecnico_laptop':
        dispositivos_permitidos = ['LAPTOP']
    else: # tecnico_general
        filtro_dispositivo = request.GET.get('filtro_dispositivo')
        if filtro_dispositivo in ['CELULAR', 'PC', 'LAPTOP']:
            dispositivos_permitidos = [filtro_dispositivo]
        else:
            dispositivos_permitidos = ['CELULAR', 'PC', 'LAPTOP']

    base_qs = Cita.objects.filter(servicio__tipo_dispositivo__in=dispositivos_permitidos)

    # Citas del técnico para la semana (incluidas las canceladas)
    citas_semana_qs = (
        base_qs
        .filter(tecnico=request.user, fecha__range=(inicio_semana, fin_semana))
        .select_related('cliente', 'servicio', 'estado')
        .order_by('fecha', 'hora_inicio')
    )

    # Si no tiene citas asignadas en esa semana, incluir todas las citas de la semana para visualización de agenda
    if not citas_semana_qs.exists():
        citas_semana_qs = (
            base_qs
            .filter(fecha__range=(inicio_semana, fin_semana))
            .select_related('cliente', 'servicio', 'estado')
            .order_by('fecha', 'hora_inicio')
        )

    citas_semana = list(citas_semana_qs)

    NOMBRES_ABREV = ['LUN', 'MAR', 'MIÉ', 'JUE', 'VIE', 'SÁB', 'DOM']
    NOMBRES_CORTOS = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    NOMBRES_COMPLETOS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    MESES = ['', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    MESES_FULL = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

    dias_semana = []
    for i in range(7):
        dia_fecha = inicio_semana + timedelta(days=i)
        citas_dia = [c for c in citas_semana if c.fecha == dia_fecha]
        dias_semana.append({
            'index': i,
            'code': f"DIA_{i}",
            'nombre_abrev': NOMBRES_ABREV[i],
            'nombre_corto': NOMBRES_CORTOS[i],
            'nombre_completo': NOMBRES_COMPLETOS[i],
            'fecha': dia_fecha,
            'fecha_str': dia_fecha.strftime('%Y-%m-%d'),
            'fecha_display': f"{NOMBRES_COMPLETOS[i]} {dia_fecha.day} de {MESES_FULL[dia_fecha.month]}",
            'dia_num': dia_fecha.day,
            'mes_num': dia_fecha.month,
            'mes_abrev': MESES[dia_fecha.month],
            'es_hoy': dia_fecha == hoy,
            'citas': citas_dia,
            'total_citas': len(citas_dia),
        })

    if inicio_semana.month == fin_semana.month:
        rango_str = f"Semana del {inicio_semana.day} al {fin_semana.day} de {MESES_FULL[inicio_semana.month]} {inicio_semana.year}"
    else:
        rango_str = f"Semana del {inicio_semana.day} de {MESES[inicio_semana.month]} al {fin_semana.day} de {MESES[fin_semana.month]} {fin_semana.year}"

    horas_grid = [f"{h:02d}:00" for h in range(8, 19)]

    # Pre-computar grilla hora x día para el calendario desktop
    # grid_rows: lista de {'hora': '08:00', 'celdas': [{'dia_index': 0, 'citas': [...], 'es_hoy': bool}]}
    grid_rows = []
    for h in range(8, 19):
        hora_str = f"{h:02d}:00"
        celdas = []
        for dia in dias_semana:
            citas_en_hora = [
                c for c in dia['citas']
                if c.hora_inicio.hour == h
            ]
            celdas.append({
                'dia_index': dia['index'],
                'es_hoy': dia['es_hoy'],
                'citas': citas_en_hora,
            })
        grid_rows.append({'hora': hora_str, 'celdas': celdas})

    from ..models import Repuesto, CitaRepuesto
    repuestos = Repuesto.objects.filter(activo=True, stock__gt=0).order_by('nombre')

    # Pre-cargar repuestos usados por citas finalizadas de esta semana
    citas_ids = [c.id for c in citas_semana]
    repuestos_por_cita = {}
    for cr in CitaRepuesto.objects.filter(cita_id__in=citas_ids).select_related('repuesto'):
        repuestos_por_cita.setdefault(cr.cita_id, []).append({
            'nombre': cr.repuesto.nombre,
            'cantidad': cr.cantidad
        })

    import json
    repuestos_por_cita_json = json.dumps({str(k): v for k, v in repuestos_por_cita.items()})

    context = {
        'dias_semana': dias_semana,
        'citas_semana': citas_semana,
        'total_citas_semana': len(citas_semana),
        'semana_offset': semana_offset,
        'semana_offset_prev': semana_offset - 1,
        'semana_offset_next': semana_offset + 1,
        'rango_str': rango_str,
        'inicio_semana': inicio_semana,
        'fin_semana': fin_semana,
        'horas_grid': horas_grid,
        'grid_rows': grid_rows,
        'hoy': hoy,
        'repuestos': repuestos,
        'repuestos_por_cita': repuestos_por_cita,
        'repuestos_por_cita_json': repuestos_por_cita_json,
    }
    return render(request, 'turnos/tecnico/tecnico_agenda.html', context)

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
@require_POST
def aceptar_cita(request, cita_id):
    """Permite al técnico aceptar una cita y pasarla a EN REPARACION."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return JsonResponse({'success': False, 'error': 'No autorizado.'}, status=403)

    cita = get_object_or_404(Cita, id=cita_id)
    
    # Check if someone else owns it
    if cita.tecnico is not None and cita.tecnico != request.user:
        return JsonResponse({'success': False, 'error': 'Esta cita ya tiene un técnico asignado.'}, status=400)

    # Asignar técnico (si no lo tenía)
    cita.tecnico = request.user
    
    # Cambiar estado a EN REPARACION
    from turnos.models.citas import EstadoCita
    estado_rep, _ = EstadoCita.objects.get_or_create(nombre='EN_REPARACION')
    cita.estado = estado_rep
    
    cita.save()

    return JsonResponse({
        'success': True,
        'message': f'Cita de {cita.cliente.nombre_completo} iniciada con éxito.'
    })

import json
from django.db import transaction

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
@require_POST
def finalizar_cita(request, cita_id):
    """Permite al técnico marcar una cita como finalizada."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return JsonResponse({'success': False, 'error': 'No autorizado.'}, status=403)

    cita = get_object_or_404(Cita, id=cita_id, tecnico=request.user)

    import json
    from ..models import CitaRepuesto

    try:
        body = json.loads(request.body)
        repuestos = body.get('repuestos', [])
    except Exception:
        repuestos = []

    # Guardar repuestos y descontar stock
    for item in repuestos:
        try:
            repuesto = Repuesto.objects.get(id=item['id'])
            cantidad = int(item['cantidad'])
            if cantidad > 0 and repuesto.stock >= cantidad:
                CitaRepuesto.objects.update_or_create(
                    cita=cita, repuesto=repuesto,
                    defaults={'cantidad': cantidad}
                )
                repuesto.stock -= cantidad
                repuesto.save()
        except (Repuesto.DoesNotExist, KeyError, ValueError):
            continue

    estado_finalizada, _ = EstadoCita.objects.get_or_create(nombre='FINALIZADA')
    cita.estado = estado_finalizada
    cita.save()

    return JsonResponse({
        'success': True,
        'message': f'Cita de {cita.cliente.nombre_completo} finalizada con éxito.'
    })

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_dispositivos(request):
    """Gestión de dispositivos e inventario para el técnico."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    # Si no hay repuestos registrados, creamos algunos datos iniciales (semilla integrada)
    if Repuesto.objects.count() == 0:
        initial_repuestos = [
            ("Pantalla iPhone 14", "OLED Original Apple", "PANTALLAS", 12, 289.00),
            ("Batería Samsung S22", "3700mAh Li-ion", "BATERIAS", 2, 45.50),
            ("Puerto Carga MacBook Pro", "MagSafe 3 Assembly", "CONECTORES", 5, 112.00),
            ("Módulo Cámara Pixel 7", "Sensor Principal 50MP", "CAMARAS", 0, 98.00),
        ]
        for name, desc, cat, stock, price in initial_repuestos:
            Repuesto.objects.create(nombre=name, descripcion=desc, categoria=cat, stock=stock, precio=price)

    repuestos = Repuesto.objects.filter(activo=True).order_by('-fecha_registro')
    
    from turnos.models.inventario import Inventario
    # Obtener la actividad reciente (salidas de inventario del usuario actual)
    actividad_reciente = Inventario.objects.filter(usuario=request.user, tipo='SALIDA').order_by('-fecha')[:5]
    
    return render(request, 'turnos/tecnico/tecnico_dispositivos.html', {
        'repuestos': repuestos,
        'actividad_reciente': actividad_reciente
    })

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_clientes(request):
    """Directorio de clientes para el técnico: muestra todos los clientes registrados."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    from django.db.models import Count, Max, Subquery, OuterRef
    from turnos.models.citas import Cita

    # Subquery: nombre del último servicio de cada cliente
    ultimo_servicio_sq = (
        Cita.objects
        .filter(cliente=OuterRef('pk'))
        .order_by('-fecha', '-hora_inicio')
        .values('servicio__nombre')[:1]
    )

    clientes = (
        Usuario.objects
        .filter(rol=Usuario.Rol.CLIENTE)
        .annotate(
            total_citas=Count('citas_cliente', distinct=True),
            ultima_fecha=Max('citas_cliente__fecha'),
            ultimo_servicio=Subquery(ultimo_servicio_sq),
        )
        .order_by('-total_citas', 'nombre_completo')
    )

    # Clasificación de nivel según cantidad de citas
    def nivel_cliente(citas):
        if citas >= 8:
            return {'label': 'Platinum', 'bg': 'bg-purple-100', 'text': 'text-purple-700'}
        elif citas >= 4:
            return {'label': 'Gold', 'bg': 'bg-yellow-100', 'text': 'text-yellow-700'}
        elif citas >= 2:
            return {'label': 'Silver', 'bg': 'bg-slate-100', 'text': 'text-slate-600'}
        else:
            return {'label': 'Nuevo', 'bg': 'bg-emerald-100', 'text': 'text-emerald-700'}

    # Agregar nivel a cada cliente (como atributo dinámico)
    clientes_con_nivel = []
    for c in clientes:
        c.nivel = nivel_cliente(c.total_citas)
        c.tag_filtro = _tag_filtro(c)
        clientes_con_nivel.append(c)

    total_clientes = len(clientes_con_nivel)
    nuevos_mes = sum(
        1 for c in clientes_con_nivel
        if c.fecha_registro and c.fecha_registro.month == __import__('datetime').date.today().month
    )

    context = {
        'clientes': clientes_con_nivel,
        'total_clientes': total_clientes,
        'nuevos_mes': nuevos_mes,
    }
    return render(request, 'turnos/tecnico/tecnico_clientes.html', context)


def _tag_filtro(cliente):
    """Genera las etiquetas de filtro (recientes, fieles, inactivos) para data-tag."""
    import datetime
    tags = []
    hoy = datetime.date.today()

    if cliente.ultima_fecha:
        dias = (hoy - cliente.ultima_fecha).days
        if dias <= 30:
            tags.append('recientes')
        elif dias > 90:
            tags.append('inactivos')

    if cliente.total_citas >= 3:
        tags.append('fieles')

    return ' '.join(tags) if tags else 'todos'



from turnos.models.soporte import TicketSoporte, EstadoSistema
import json
from django.http import JsonResponse
from django.views.decorators.http import require_POST

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_soporte(request):
    """Soporte operativo para el técnico."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')
        
    tickets = TicketSoporte.objects.filter(tecnico=request.user).order_by('-fecha_creacion')
    sistemas = EstadoSistema.objects.all().order_by('nombre')
    
    context = {
        'tickets': tickets,
        'sistemas': sistemas
    }
    return render(request, 'turnos/tecnico/tecnico_soporte.html', context)

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
@require_POST
def tecnico_crear_ticket(request):
    if request.user.rol != Usuario.Rol.TECNICO:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        data = json.loads(request.body)
        titulo = data.get('titulo')
        area = data.get('area')
        urgencia = data.get('urgencia')
        descripcion = data.get('descripcion')
        
        ticket = TicketSoporte.objects.create(
            titulo=titulo,
            area=area,
            urgencia=urgencia,
            descripcion=descripcion,
            tecnico=request.user
        )
        return JsonResponse({'success': True, 'message': 'Ticket creado con éxito'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)



@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_reporte_mensual(request):
    """
    Genera el reporte mensual del técnico como archivo Excel (.xlsx).
    Parámetros GET: mes (1-12), anio (ej. 2026)
    """
    from datetime import date
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    hoy = date.today()
    try:
        mes  = int(request.GET.get('mes',  hoy.month))
        anio = int(request.GET.get('anio', hoy.year))
        if not (1 <= mes <= 12):  mes  = hoy.month
        if not (2000 <= anio <= 2100): anio = hoy.year
    except (ValueError, TypeError):
        mes, anio = hoy.month, hoy.year

    MESES_ES = {
        1:'Enero', 2:'Febrero', 3:'Marzo', 4:'Abril',
        5:'Mayo',  6:'Junio',   7:'Julio', 8:'Agosto',
        9:'Septiembre', 10:'Octubre', 11:'Noviembre', 12:'Diciembre',
    }
    nombre_mes = MESES_ES.get(mes, str(mes))

    citas = (
        Cita.objects
        .filter(tecnico=request.user, fecha__year=anio, fecha__month=mes)
        .select_related('cliente', 'servicio', 'estado')
        .order_by('fecha', 'hora_inicio')
    )

    total      = citas.count()
    finalizadas = citas.filter(estado__nombre__iexact='FINALIZADA').count()
    canceladas  = citas.filter(estado__nombre__iexact='CANCELADA').count()
    pendientes  = citas.exclude(estado__nombre__in=['FINALIZADA', 'CANCELADA']).count()
    tasa        = f"{round(finalizadas / total * 100, 1)}%" if total > 0 else "0%"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Servicios"

    # Habilitar cuadrícula visible
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_section = Font(name='Arial', size=11, bold=True, color='002B75')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_body = Font(name='Arial', size=10)
    font_bold = Font(name='Arial', size=10, bold=True)
    
    fill_navy = PatternFill(start_color='002B75', end_color='002B75', fill_type='solid')
    fill_light_blue = PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid')
    fill_gray = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Título / Banner
    ws.merge_cells('A1:H2')
    title_cell = ws['A1']
    title_cell.value = f"REPORTE MENSUAL DE SERVICIOS - {nombre_mes.upper()} {anio}"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = align_center

    # 2. Información del Técnico / Generación
    ws['A4'] = "Técnico:"
    ws['A4'].font = font_bold
    ws['B4'] = request.user.nombre_completo
    ws['B4'].font = font_body

    ws['A5'] = "Correo:"
    ws['A5'].font = font_bold
    ws['B5'] = request.user.correo
    ws['B5'].font = font_body

    ws['G4'] = "Generación:"
    ws['G4'].font = font_bold
    ws['H4'] = hoy.strftime('%d/%m/%Y')
    ws['H4'].font = font_body

    # 3. KPIs
    ws['A7'] = "KPIs del Periodo"
    ws['A7'].font = font_section

    ws['A8'] = "Métrica"
    ws['A8'].font = font_header
    ws['A8'].fill = fill_navy
    ws['A8'].alignment = align_center
    ws['B8'] = "Valor"
    ws['B8'].font = font_header
    ws['B8'].fill = fill_navy
    ws['B8'].alignment = align_center

    metrics = [
        ("Total Citas", total),
        ("Finalizadas", finalizadas),
        ("Canceladas", canceladas),
        ("Pendientes", pendientes),
        ("Tasa Completado", tasa)
    ]
    for idx, (m_name, m_val) in enumerate(metrics, start=9):
        ws[f'A{idx}'] = m_name
        ws[f'A{idx}'].font = font_bold
        ws[f'A{idx}'].border = border_all
        ws[f'A{idx}'].fill = fill_gray
        
        ws[f'B{idx}'] = m_val
        ws[f'B{idx}'].font = font_body
        ws[f'B{idx}'].border = border_all
        ws[f'B{idx}'].alignment = align_center

    # 4. Tabla de Citas
    ws['A15'] = "Detalle de Citas"
    ws['A15'].font = font_section

    headers = ["#", "Fecha", "Horario", "Cliente", "Servicio", "Estado", "Retraso", "Observaciones"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=17, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    row_num = 18
    if total == 0:
        ws.merge_cells('A18:H18')
        empty_cell = ws['A18']
        empty_cell.value = "Sin citas registradas en este período."
        empty_cell.font = font_body
        empty_cell.alignment = align_center
        for col_idx in range(1, 9):
            ws.cell(row=18, column=col_idx).border = border_all
    else:
        for idx, c in enumerate(citas, start=1):
            ws.cell(row=row_num, column=1, value=f"{idx:02d}").alignment = align_center
            ws.cell(row=row_num, column=2, value=c.fecha.strftime('%d/%m/%Y')).alignment = align_center
            ws.cell(row=row_num, column=3, value=f"{c.hora_inicio.strftime('%H:%M')} – {c.hora_fin.strftime('%H:%M')}").alignment = align_center
            ws.cell(row=row_num, column=4, value=c.cliente.nombre_completo)
            ws.cell(row=row_num, column=5, value=c.servicio.nombre if c.servicio else '—')
            ws.cell(row=row_num, column=6, value=c.estado.nombre if c.estado else '—').alignment = align_center
            ws.cell(row=row_num, column=7, value=f"{c.minutos_retraso} min" if c.minutos_retraso > 0 else '—').alignment = align_center
            ws.cell(row=row_num, column=8, value=c.observaciones if c.observaciones else '—')

            for col_idx in range(1, 9):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = font_body
                cell.border = border_all
                if idx % 2 == 0:
                    cell.fill = fill_gray
            row_num += 1

    # Ajustar ancho de columnas (evita MergedCell sin column_letter)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 9):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Reporte_Servicios_{mes}_{anio}.xlsx'
    wb.save(response)
    return response



@login_required
@rol_requerido([Usuario.Rol.CLIENTE])
def cliente_inicio(request):
    """Dashboard principal del cliente (Inicio)"""
    if request.user.rol != Usuario.Rol.CLIENTE:
        return redirect('home')

    from turnos.services.cancelacion_automatica import cancelar_citas_vencidas
    cancelar_citas_vencidas()

    citas_usuario = Cita.objects.filter(cliente=request.user).order_by('-fecha', '-hora_inicio')

    total_citas_activas = citas_usuario.exclude(estado__nombre__in=['FINALIZADA', 'CANCELADA']).count()
    total_reparaciones = citas_usuario.filter(estado__nombre='FINALIZADA').count()

    citas_recientes = citas_usuario

    from turnos.models.notificaciones import Notificacion
    notificaciones = Notificacion.objects.filter(usuario=request.user, leida=False).order_by('-fecha_envio')

    # Citas canceladas recientemente por inasistencia para alerta visual
    citas_canceladas_recientes = citas_usuario.filter(
        estado__nombre='CANCELADA',
        observaciones__icontains='inasistencia'
    )[:3]

    context = {
        'total_citas_activas': total_citas_activas,
        'total_reparaciones': total_reparaciones,
        'citas_recientes': citas_recientes,
        'notificaciones': notificaciones,
        'citas_canceladas_recientes': citas_canceladas_recientes,
    }
    return render(request, 'turnos/cliente/cliente_inicio.html', context)

@login_required
@rol_requerido([Usuario.Rol.CLIENTE])
def cliente_servicios(request):
    """Catálogo de servicios para el cliente"""
    if request.user.rol != Usuario.Rol.CLIENTE:
        return redirect('home')
    
    from ..models import Servicio
    servicios = Servicio.objects.filter(activo=True).order_by('nombre')
    return render(request, 'turnos/cliente/cliente_servicios.html', {'servicios': servicios})

@login_required
@rol_requerido([Usuario.Rol.CLIENTE])
def cliente_perfil(request):
    """Perfil del cliente: permite editar datos personales."""
    if request.user.rol != Usuario.Rol.CLIENTE:
        return redirect('home')

    if request.method == 'POST':
        form = EditarPerfilForm(
            request.POST,
            instance=request.user,
            current_user=request.user,
        )
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tus datos personales fueron actualizados correctamente!')
            return redirect('cliente_perfil')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = EditarPerfilForm(instance=request.user, current_user=request.user)

    return render(request, 'turnos/cliente/cliente_perfil.html', {'form': form})

@login_required
@rol_requerido([Usuario.Rol.CLIENTE])
def cliente_soporte(request):
    """Página de soporte para el cliente"""
    if request.user.rol != Usuario.Rol.CLIENTE:
        return redirect('home')
    return render(request, 'turnos/cliente/cliente_soporte.html')
@login_required
@rol_requerido([Usuario.Rol.CLIENTE])
def cliente_notificaciones(request):
    """Módulo de notificaciones para el cliente."""
    if request.user.rol != Usuario.Rol.CLIENTE:
        return redirect('home')
    
    from turnos.models.notificaciones import Notificacion
    notificaciones = Notificacion.objects.filter(usuario=request.user).order_by('-fecha_envio')
    
    # Marcar como leídas automáticamente al entrar a la vista (opcional pero recomendado)
    notifs_no_leidas = notificaciones.filter(leida=False)
    if notifs_no_leidas.exists():
        notifs_no_leidas.update(leida=True)
        
    return render(request, 'turnos/cliente/cliente_notificaciones.html', {
        'notificaciones': notificaciones
    })


@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_perfil(request):
    """Perfil del técnico para ver métricas y actualizar información personal."""
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    if request.method == 'POST':
        form = EditarPerfilForm(request.POST, instance=request.user, current_user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu perfil técnico fue actualizado con éxito!')
            return redirect('tecnico_perfil')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = EditarPerfilForm(instance=request.user, current_user=request.user)

    return render(request, 'turnos/tecnico/tecnico_perfil.html', {'form': form})


@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def exportar_inventario_excel(request):
    """Exporta la lista de repuestos del técnico a un archivo Excel (.xlsx)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    if request.user.rol not in [Usuario.Rol.TECNICO, Usuario.Rol.ADMINISTRADOR]:
        return redirect('home')

    repuestos = Repuesto.objects.filter(activo=True).order_by('categoria', 'nombre')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario Repuestos"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_body = Font(name='Arial', size=10)
    font_bold = Font(name='Arial', size=10, bold=True)
    
    fill_navy = PatternFill(start_color='002B75', end_color='002B75', fill_type='solid')
    fill_gray = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')
    
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Título
    ws.merge_cells('A1:E2')
    title_cell = ws['A1']
    title_cell.value = "INVENTARIO DE REPUESTOS Y COMPONENTES"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = align_center

    # Headers
    headers = ["Componente", "Categoría", "Stock (Unidades)", "Precio Unitario", "Estado"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    row_num = 5
    for idx, r in enumerate(repuestos, start=1):
        ws.cell(row=row_num, column=1, value=r.nombre)
        ws.cell(row=row_num, column=2, value=r.get_categoria_display()).alignment = align_center
        ws.cell(row=row_num, column=3, value=r.stock).alignment = align_center
        
        cell_precio = ws.cell(row=row_num, column=4, value=float(r.precio))
        cell_precio.number_format = '$#,##0.00'
        cell_precio.alignment = align_right
        
        estado = "Disponible" if r.stock > 5 else ("Stock Bajo" if r.stock > 0 else "Agotado")
        ws.cell(row=row_num, column=5, value=estado).alignment = align_center

        for col_idx in range(1, 6):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_body
            cell.border = border_all
            if idx % 2 == 0:
                cell.fill = fill_gray
        row_num += 1

    # Ajustar columnas (evita MergedCell sin column_letter)
    from openpyxl.utils import get_column_letter
    for col_idx in range(1, 6):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_Repuestos.xlsx'
    wb.save(response)
    return response


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_usuarios(request):
    if not request.user.es_admin:
        return redirect('home')
    
    hoy = date.today()
    inicio_mes = date(hoy.year, hoy.month, 1)

    qs = Usuario.objects.all().order_by('-fecha_registro')

    # Filtros desde GET
    search = request.GET.get('search', '').strip()
    rol = request.GET.get('rol', '').strip()
    estado = request.GET.get('estado', '').strip()

    if search:
        from django.db.models import Q
        qs = qs.filter(Q(nombre_completo__icontains=search) | Q(correo__icontains=search) | Q(telefono__icontains=search))
    if rol and rol != 'TODOS':
        qs = qs.filter(rol=rol)
    if estado:
        if estado == 'ACTIVO':
            qs = qs.filter(activo=True)
        elif estado == 'INACTIVO':
            qs = qs.filter(activo=False)

    # Conteo general para KPIs
    total_usuarios   = Usuario.objects.count()
    usuarios_activos = Usuario.objects.filter(activo=True).count()
    nuevos_mes       = Usuario.objects.filter(fecha_registro__gte=inicio_mes).count()

    # Conteo por rol
    total_clientes = Usuario.objects.filter(rol=Usuario.Rol.CLIENTE).count()
    total_tecnicos = Usuario.objects.filter(rol=Usuario.Rol.TECNICO).count()
    total_admins   = Usuario.objects.filter(rol=Usuario.Rol.ADMINISTRADOR).count()

    # Paginación (10 usuarios por página)
    from django.core.paginator import Paginator
    paginator   = Paginator(qs, 10)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    context = {
        'usuarios':         page_obj,
        'page_obj':         page_obj,
        'total_filtrados':  qs.count(),
        'total_usuarios':   total_usuarios,
        'usuarios_activos': usuarios_activos,
        'nuevos_mes':       nuevos_mes,
        'total_clientes':   total_clientes,
        'total_tecnicos':   total_tecnicos,
        'total_admins':     total_admins,
        'search':           search,
        'filtro_rol':       rol,
        'filtro_estado':    estado,
    }
    return render(request, 'turnos/administracion/admin_usuarios.html', context)


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_crear_usuario(request):
    if not request.user.es_admin:
        return redirect('home')
    if request.method == 'POST':
        nombre   = request.POST.get('nombre_completo', '').strip()
        correo   = request.POST.get('correo', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        rol      = request.POST.get('rol', Usuario.Rol.CLIENTE)
        password = request.POST.get('password', '')
        if not nombre or not correo or not password:
            messages.error(request, 'Nombre, correo y contraseña son obligatorios.')
        elif Usuario.objects.filter(correo=correo).exists():
            messages.error(request, f'Ya existe un usuario con el correo {correo}.')
        else:
            try:
                Usuario.objects.create_user(correo=correo, password=password,
                    nombre_completo=nombre, telefono=telefono or None, rol=rol)
                messages.success(request, f'Usuario "{nombre}" creado correctamente.')
            except Exception as e:
                messages.error(request, f'Error: {e}')
    return redirect('admin_usuarios')


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_editar_usuario(request, usuario_id):
    if not request.user.es_admin:
        return redirect('home')
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    if request.method == 'POST':
        usuario.nombre_completo = request.POST.get('nombre_completo', usuario.nombre_completo).strip()
        usuario.correo          = request.POST.get('correo', usuario.correo).strip()
        usuario.telefono        = request.POST.get('telefono', '').strip() or None
        usuario.rol             = request.POST.get('rol', usuario.rol)
        try:
            usuario.save()
            messages.success(request, f'Usuario "{usuario.nombre_completo}" actualizado.')
        except Exception as e:
            messages.error(request, f'Error: {e}')
    return redirect('admin_usuarios')


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_toggle_usuario(request, usuario_id):
    if not request.user.es_admin:
        return redirect('home')
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    if request.method == 'POST':
        if usuario.pk == request.user.pk:
            messages.error(request, 'No puedes desactivar tu propia cuenta.')
        else:
            usuario.activo = not usuario.activo
            usuario.save()
            messages.success(request, f'Usuario "{usuario.nombre_completo}" {"activado" if usuario.activo else "desactivado"}.')
    return redirect('admin_usuarios')


# ─────────────────────────────────────────────
#  CATÁLOGO DE SERVICIOS (ADMIN)
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_servicios(request):
    if not request.user.es_admin:
        return redirect('home')
    servicios      = Servicio.objects.select_related('especialidad').order_by('tipo_dispositivo', 'nombre')
    especialidades = Especialidad.objects.filter(activo=True).order_by('nombre')

    # Búsqueda global desde topbar
    search_query = request.GET.get('search', '').strip()
    if search_query:
        from django.db.models import Q
        servicios = servicios.filter(
            Q(nombre__icontains=search_query) |
            Q(descripcion__icontains=search_query) |
            Q(tipo_dispositivo__icontains=search_query)
        )

    servicios_activos   = Servicio.objects.filter(activo=True).count()
    servicios_inactivos = Servicio.objects.filter(activo=False).count()
    servicios_premium   = Servicio.objects.filter(activo=True).order_by('tipo_dispositivo', 'nombre')[:4]
    return render(request, 'turnos/administracion/admin_servicios.html', {
        'servicios':           servicios,
        'especialidades':      especialidades,
        'servicios_activos':   servicios_activos,
        'servicios_inactivos': servicios_inactivos,
        'servicios_premium':   servicios_premium,
        'search_query':        search_query,
    })


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_crear_servicio(request):
    if not request.user.es_admin:
        return redirect('home')
    if request.method == 'POST':
        nombre      = request.POST.get('nombre', '').strip()
        desc        = request.POST.get('descripcion', '').strip()
        tipo        = request.POST.get('tipo_dispositivo', 'CELULAR')
        duracion    = int(request.POST.get('duracion_minutos', 60) or 60)
        buffer      = int(request.POST.get('buffer_minutos', 0) or 0)
        esp_id      = request.POST.get('especialidad_id')
        
        if not nombre:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            try:
                especialidad = None
                if esp_id:
                    especialidad = Especialidad.objects.filter(pk=esp_id).first()
                if not especialidad:
                    especialidad, _ = Especialidad.objects.get_or_create(nombre='HARDWARE', defaults={'descripcion': 'General Hardware'})

                Servicio.objects.create(
                    nombre=nombre,
                    descripcion=desc or None,
                    tipo_dispositivo=tipo,
                    duracion_minutos=duracion,
                    buffer_minutos=buffer,
                    especialidad=especialidad
                )
                messages.success(request, f'¡Servicio "{nombre}" creado con éxito!')
            except Exception as e:
                messages.error(request, f'Error al crear servicio: {e}')
    return redirect('admin_servicios')


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_editar_servicio(request, servicio_id):
    if not request.user.es_admin:
        return redirect('home')
    servicio = get_object_or_404(Servicio, pk=servicio_id)
    if request.method == 'POST':
        servicio.nombre           = request.POST.get('nombre', servicio.nombre).strip()
        servicio.descripcion      = request.POST.get('descripcion', '').strip() or None
        servicio.tipo_dispositivo = request.POST.get('tipo_dispositivo', servicio.tipo_dispositivo)
        servicio.duracion_minutos = int(request.POST.get('duracion_minutos', servicio.duracion_minutos) or 60)
        servicio.buffer_minutos   = int(request.POST.get('buffer_minutos', servicio.buffer_minutos) or 0)
        esp_id = request.POST.get('especialidad_id')
        if esp_id:
            esp = Especialidad.objects.filter(pk=esp_id).first()
            if esp:
                servicio.especialidad = esp
        try:
            servicio.save()
            messages.success(request, f'¡Servicio "{servicio.nombre}" actualizado correctamente!')
        except Exception as e:
            messages.error(request, f'Error al actualizar servicio: {e}')
    return redirect('admin_servicios')


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_toggle_servicio(request, servicio_id):
    if not request.user.es_admin:
        return redirect('home')
    servicio = get_object_or_404(Servicio, pk=servicio_id)
    if request.method == 'POST':
        servicio.activo = not servicio.activo
        servicio.save()
        messages.success(request, f'Servicio "{servicio.nombre}" {"activado" if servicio.activo else "desactivado"}.')
    return redirect('admin_servicios')


# ─────────────────────────────────────────────
#  CITAS (ADMIN)
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_citas(request):
    if request.method == 'POST':
        accion = request.POST.get('accion')
        cita_id = request.POST.get('cita_id')
        if accion == 'cancelar_cita' and cita_id:
            
            try:
                cita = Cita.objects.get(id=cita_id)
                estado_cancelada, _ = EstadoCita.objects.get_or_create(nombre='CANCELADA')
                cita.estado = estado_cancelada
                cita.save()
                messages.success(request, f'La cita #{cita.id} ha sido marcada como cancelada.')
            except Cita.DoesNotExist:
                messages.error(request, 'Cita no encontrada.')
        elif accion == 'eliminar_cita' and cita_id:
            
            try:
                cita = Cita.objects.get(id=cita_id)
                cita_id_num = cita.id
                cita.delete()
                messages.success(request, f'La cita #{cita_id_num} ha sido eliminada definitivamente.')
            except Cita.DoesNotExist:
                messages.error(request, 'Cita no encontrada.')
        return redirect('admin_citas')

    hoy = date.today()

    # ── KPIs ──
    citas_hoy      = Cita.objects.filter(fecha=hoy).count()
    pendientes     = Cita.objects.filter(estado__nombre__iexact='CONFIRMADA').count()
    en_proceso     = Cita.objects.filter(estado__nombre__iexact='EN_REPARACION').count()
    completadas    = Cita.objects.filter(estado__nombre__iexact='FINALIZADA').count()

    # ── Filtros desde GET ──
    qs = Cita.objects.select_related('cliente', 'tecnico', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    filtro_cliente  = request.GET.get('cliente', '').strip()
    filtro_tecnico  = request.GET.get('tecnico', '').strip()
    filtro_estado   = request.GET.get('estado', '').strip()
    filtro_fecha    = request.GET.get('fecha', '').strip()
    filtro_fecha_ini = request.GET.get('fecha_ini', '').strip()
    filtro_fecha_fin = request.GET.get('fecha_fin', '').strip()

    if filtro_cliente:
        qs = qs.filter(cliente__nombre_completo__icontains=filtro_cliente)
    if filtro_tecnico:
        qs = qs.filter(tecnico__id=filtro_tecnico)
    if filtro_estado:
        qs = qs.filter(estado__nombre__iexact=filtro_estado)

    # Filtro rápido de fecha (Hoy / Ayer)
    from datetime import timedelta
    if filtro_fecha == 'hoy':
        qs = qs.filter(fecha=hoy)
    elif filtro_fecha == 'ayer':
        qs = qs.filter(fecha=hoy - timedelta(days=1))

    if filtro_fecha_ini:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__gte=datetime.strptime(filtro_fecha_ini, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filtro_fecha_fin:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__lte=datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Paginación simple (25 por página)
    from django.core.paginator import Paginator
    paginator   = Paginator(qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    tecnicos    = Usuario.objects.filter(rol=Usuario.Rol.TECNICO, activo=True).order_by('nombre_completo')
    # Deduplicar estados por nombre (ignorando mayúsculas/minúsculas)
    todos_estados = EstadoCita.objects.all().order_by('nombre')
    vistos = set()
    estados = []
    for est in todos_estados:
        clave = est.nombre.upper().strip()
        if clave not in vistos:
            vistos.add(clave)
            estados.append(est)

    context = {
        'citas':           page_obj,
        'page_obj':        page_obj,
        'total_citas':     qs.count(),
        'citas_hoy':       citas_hoy,
        'pendientes':      pendientes,
        'en_proceso':      en_proceso,
        'completadas':     completadas,
        'tecnicos':        tecnicos,
        'estados':         estados,
        'hoy':             hoy,
        'manana':          hoy + timedelta(days=1),
        'ayer':            hoy - timedelta(days=1),
        'filtro_cliente':  filtro_cliente,
        'filtro_tecnico':  filtro_tecnico,
        'filtro_estado':   filtro_estado,
        'filtro_fecha':    filtro_fecha,
        'filtro_fecha_ini': filtro_fecha_ini,
        'filtro_fecha_fin': filtro_fecha_fin,
    }
    return render(request, 'turnos/administracion/admin_citas.html', context)


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_exportar_citas_excel(request):
    """Exporta la lista de citas a Excel con diseño profesional."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    hoy = date.today()

    qs = Cita.objects.select_related('cliente', 'tecnico', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gestión de Citas"
    ws.sheet_view.showGridLines = False

    # Estilos
    navy_fill    = PatternFill("solid", fgColor="1D4ED8")
    blue_fill    = PatternFill("solid", fgColor="EFF6FF")
    gray_fill    = PatternFill("solid", fgColor="F8FAFC")
    green_fill   = PatternFill("solid", fgColor="D1FAE5")
    amber_fill   = PatternFill("solid", fgColor="FEF3C7")
    red_fill     = PatternFill("solid", fgColor="FEE2E2")
    purple_fill  = PatternFill("solid", fgColor="EDE9FE")

    thin = Side(border_style="thin", color="E2E8F0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    font_title   = Font(name="Inter", size=16, bold=True, color="1D4ED8")
    font_sub     = Font(name="Inter", size=10, color="64748B")
    font_header  = Font(name="Inter", size=10, bold=True, color="FFFFFF")
    font_body    = Font(name="Inter", size=9, color="1E293B")
    font_bold    = Font(name="Inter", size=9, bold=True, color="1E293B")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── Título ──
    ws.merge_cells("A1:H2")
    ws["A1"] = "ServiTech — Reporte de Gestión de Citas"
    ws["A1"].font = font_title
    ws["A1"].alignment = center
    ws["A1"].fill = blue_fill

    ws.merge_cells("A3:H3")
    ws["A3"] = f"Generado el {hoy.strftime('%d/%m/%Y')}  |  Total de registros: {qs.count()}"
    ws["A3"].font = font_sub
    ws["A3"].alignment = center

    # ── KPIs ──
    ws.merge_cells("A4:B4"); ws["A4"] = "Citas Hoy"
    ws.merge_cells("C4:D4"); ws["C4"] = "Pendientes"
    ws.merge_cells("E4:F4"); ws["E4"] = "Completadas"

    ws.merge_cells("A5:B5"); ws["A5"] = Cita.objects.filter(fecha=hoy).count()
    ws.merge_cells("C5:D5"); ws["C5"] = Cita.objects.filter(estado__nombre__iexact='CONFIRMADA').count()
    ws.merge_cells("E5:F5"); ws["E5"] = Cita.objects.filter(estado__nombre__iexact='FINALIZADA').count()

    for cell in ["A4","C4","E4"]:
        ws[cell].font = Font(name="Inter", size=9, bold=True, color="64748B")
        ws[cell].alignment = center
    for cell in ["A5","C5","E5"]:
        ws[cell].font = Font(name="Inter", size=18, bold=True, color="1D4ED8")
        ws[cell].alignment = center

    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 30

    # ── Espacio ──
    ws.row_dimensions[6].height = 8

    # ── Encabezados tabla ──
    headers = ["#", "CLIENTE", "SERVICIO", "TÉCNICO", "FECHA", "HORA", "ESTADO", "OBSERVACIONES"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = font_header
        cell.fill = navy_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[7].height = 22

    # ── Datos ──
    row_num = 8
    for idx, cita in enumerate(qs, 1):
        estado_nombre = cita.estado.nombre if cita.estado else "Sin estado"

        # Color de fila por estado
        if estado_nombre.lower() in ['finalizada']:
            row_fill = green_fill
        elif estado_nombre.lower() in ['cancelada']:
            row_fill = red_fill
        elif estado_nombre.lower() in ['diagnóstico', 'diagnostico', 'en_reparacion']:
            row_fill = purple_fill
        elif estado_nombre.lower() in ['confirmada']:
            row_fill = amber_fill
        else:
            row_fill = gray_fill if idx % 2 == 0 else None

        data = [
            f"{idx:03d}",
            cita.cliente.nombre_completo if cita.cliente else "—",
            cita.servicio.nombre if cita.servicio else "—",
            cita.tecnico.nombre_completo if cita.tecnico else "Sin asignar",
            cita.fecha.strftime('%d/%m/%Y') if cita.fecha else "—",
            f"{cita.hora_inicio.strftime('%H:%M')} — {cita.hora_fin.strftime('%H:%M')}",
            estado_nombre,
            cita.observaciones or "—",
        ]

        aligns = [center, left, left, left, center, center, center, left]

        for col, (val, aln) in enumerate(zip(data, aligns), 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = font_bold if col == 1 else font_body
            cell.alignment = aln
            cell.border = border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_num].height = 18
        row_num += 1

    # ── Anchos de columna ──
    col_widths = [6, 28, 28, 24, 14, 18, 16, 35]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Respuesta ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=ServiTech_Citas_{hoy.strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
#  REPORTES (ADMIN)
# ─────────────────────────────────────────────
@login_required
def admin_reportes(request):
    if not request.user.es_admin:
        return redirect('home')
    hoy = date.today()
    citas_mes_qs = Cita.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
    total       = citas_mes_qs.count()
    finalizadas = citas_mes_qs.filter(estado__nombre__iexact='FINALIZADA').count()
    canceladas  = citas_mes_qs.filter(estado__nombre__iexact='CANCELADA').count()
    en_proceso  = citas_mes_qs.filter(estado__nombre__in=['EN_REPARACION']).count()
    tasa_exito  = round(finalizadas / total * 100, 1) if total > 0 else 0.0

    # Servicios populares
    servicios_populares = (Cita.objects
        .filter(fecha__year=hoy.year, fecha__month=hoy.month)
        .values('servicio__nombre').annotate(total=Count('id')).order_by('-total')[:5])

    # Distribución por tipo de dispositivo
    dist_dispositivos = (Cita.objects
        .values('servicio__tipo_dispositivo')
        .annotate(total=Count('id'))
        .order_by('-total'))
    
    total_dispositivos = sum(item['total'] for item in dist_dispositivos) or 1
    tipos_stats = {
        'celulares_pct': round(sum(item['total'] for item in dist_dispositivos if item['servicio__tipo_dispositivo'] == 'CELULAR') / total_dispositivos * 100, 1),
        'laptops_pct':   round(sum(item['total'] for item in dist_dispositivos if item['servicio__tipo_dispositivo'] == 'LAPTOP') / total_dispositivos * 100, 1),
        'pc_pct':        round(sum(item['total'] for item in dist_dispositivos if item['servicio__tipo_dispositivo'] == 'PC') / total_dispositivos * 100, 1),
    }

    # Técnicos TOP
    tecnicos_top = (Cita.objects
        .filter(fecha__year=hoy.year, fecha__month=hoy.month, tecnico__isnull=False)
        .values('tecnico__nombre_completo').annotate(total=Count('id')).order_by('-total')[:5])

    # Tendencia de los últimos 6 meses para el gráfico Chart.js
    MESES_ES = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    chart_labels = []
    chart_data = []
    chart_metas = []
    for i in range(5, -1, -1):
        primer_dia = date(hoy.year, hoy.month, 1)
        mes_obj = primer_dia
        for _ in range(i):
            mes_obj = (mes_obj.replace(day=1) - timedelta(days=1)).replace(day=1)
        ultimo = calendar.monthrange(mes_obj.year, mes_obj.month)[1]
        cnt = Cita.objects.filter(fecha__gte=mes_obj, fecha__lte=mes_obj.replace(day=ultimo), estado__nombre__iexact='FINALIZADA').count()
        chart_labels.append(MESES_ES[mes_obj.month - 1])
        chart_data.append(cnt)
        chart_metas.append(max(cnt + 2, 5))

    # Citas/reportes recientes (últimos 5 registros)
    reportes_recientes = Cita.objects.select_related('cliente', 'servicio', 'estado', 'tecnico').order_by('-fecha_creacion')[:5]

    context = {
        'reporte': {
            'total': total,
            'finalizadas': finalizadas,
            'canceladas': canceladas,
            'en_proceso': en_proceso,
            'tasa_exito': tasa_exito
        },
        'servicios_populares': servicios_populares,
        'tipos_stats':         tipos_stats,
        'tecnicos_top':        tecnicos_top,
        'chart_labels_json':   chart_labels,
        'chart_data_json':     chart_data,
        'chart_metas_json':    chart_metas,
        'reportes_recientes':  reportes_recientes,
    }
    return render(request, 'turnos/administracion/admin_reportes.html', context)


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_exportar_analitico_pdf(request):
    """Exporta el informe analítico completo a PDF con reportlab."""
    if not request.user.es_admin:
        return redirect('home')

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import io

    hoy = date.today()
    citas_mes_qs = Cita.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
    total       = citas_mes_qs.count()
    finalizadas = citas_mes_qs.filter(estado__nombre__iexact='FINALIZADA').count()
    canceladas  = citas_mes_qs.filter(estado__nombre__iexact='CANCELADA').count()
    en_proceso  = citas_mes_qs.filter(estado__nombre__in=['EN_REPARACION']).count()
    tasa_exito  = round(finalizadas / total * 100, 1) if total > 0 else 0.0

    reportes_recientes = Cita.objects.select_related('cliente','servicio','estado','tecnico').order_by('-fecha_creacion')[:10]

    servicios_populares = (Cita.objects
        .filter(fecha__year=hoy.year, fecha__month=hoy.month)
        .values('servicio__nombre').annotate(total=Count('id')).order_by('-total')[:5])

    tecnicos_top = (Cita.objects
        .filter(fecha__year=hoy.year, fecha__month=hoy.month, tecnico__isnull=False)
        .values('tecnico__nombre_completo').annotate(total=Count('id')).order_by('-total')[:5])

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.8*cm, leftMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm,
        title='ServiTech — Informe Analítico'
    )

    AZUL      = colors.HexColor('#003399')
    AZUL_CLARO= colors.HexColor('#EEF2FF')
    GRIS      = colors.HexColor('#64748b')
    GRIS_BORDE= colors.HexColor('#E2E8F0')
    VERDE     = colors.HexColor('#047857')
    VERDE_BG  = colors.HexColor('#D1FAE5')
    BLANCO    = colors.white

    estilos = getSampleStyleSheet()
    titulo_e   = ParagraphStyle('titulo',   fontSize=18, textColor=AZUL,  fontName='Helvetica-Bold', spaceAfter=2)
    subtitulo_e= ParagraphStyle('sub',      fontSize=9,  textColor=GRIS,  fontName='Helvetica',      spaceAfter=6)
    seccion_e  = ParagraphStyle('seccion',  fontSize=11, textColor=AZUL,  fontName='Helvetica-Bold', spaceBefore=10, spaceAfter=4)
    pie_e      = ParagraphStyle('pie',      fontSize=7,  textColor=GRIS,  fontName='Helvetica',      alignment=TA_RIGHT)

    elementos = []

    # ── Encabezado ──────────────────────────────
    elementos.append(Paragraph('ServiTech — Informe Analítico de Operaciones', titulo_e))
    elementos.append(Paragraph(f'Generado el {hoy.strftime("%d/%m/%Y")} · Datos del mes en curso', subtitulo_e))
    elementos.append(HRFlowable(width='100%', thickness=1.5, color=AZUL, spaceAfter=10))

    # ── KPIs ────────────────────────────────────
    elementos.append(Paragraph('Resumen del Mes', seccion_e))
    kpi_data = [
        ['Citas Totales', 'Finalizadas', 'Canceladas', 'En Proceso', 'Tasa de Éxito'],
        [str(total), str(finalizadas), str(canceladas), str(en_proceso), f'{tasa_exito}%'],
    ]
    kpi_tabla = Table(kpi_data, colWidths=[3.2*cm]*5)
    kpi_tabla.setStyle(TableStyle([
        ('BACKGROUND',  (0,0), (-1,0), AZUL),
        ('TEXTCOLOR',   (0,0), (-1,0), BLANCO),
        ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,0), (-1,0), 8),
        ('BACKGROUND',  (0,1), (-1,1), AZUL_CLARO),
        ('FONTNAME',    (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE',    (0,1), (-1,1), 14),
        ('TEXTCOLOR',   (0,1), (-1,1), AZUL),
        ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
        ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUND',(0,1),(-1,1), AZUL_CLARO),
        ('GRID',        (0,0), (-1,-1), 0.5, GRIS_BORDE),
        ('TOPPADDING',  (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
        ('ROUNDEDCORNERS', [4]),
    ]))
    elementos.append(kpi_tabla)
    elementos.append(Spacer(1, 10))

    # ── Servicios populares ──────────────────────
    if servicios_populares:
        elementos.append(Paragraph('Servicios Más Solicitados', seccion_e))
        sp_data = [['Servicio', 'Cantidad']] + [[s['servicio__nombre'], str(s['total'])] for s in servicios_populares]
        sp_tabla = Table(sp_data, colWidths=[12*cm, 4*cm])
        sp_tabla.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), AZUL),
            ('TEXTCOLOR',     (0,0), (-1,0), BLANCO),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [BLANCO, AZUL_CLARO]),
            ('ALIGN',         (1,0), (1,-1), 'CENTER'),
            ('GRID',          (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elementos.append(sp_tabla)
        elementos.append(Spacer(1, 10))

    # ── Técnicos top ────────────────────────────
    if tecnicos_top:
        elementos.append(Paragraph('Técnicos con Más Atenciones', seccion_e))
        tt_data = [['Técnico', 'Atenciones']] + [[t['tecnico__nombre_completo'], str(t['total'])] for t in tecnicos_top]
        tt_tabla = Table(tt_data, colWidths=[12*cm, 4*cm])
        tt_tabla.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0), AZUL),
            ('TEXTCOLOR',     (0,0), (-1,0), BLANCO),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [BLANCO, AZUL_CLARO]),
            ('ALIGN',         (1,0), (1,-1), 'CENTER'),
            ('GRID',          (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ('TOPPADDING',    (0,0), (-1,-1), 6),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        elementos.append(tt_tabla)
        elementos.append(Spacer(1, 10))

    # ── Atenciones recientes ─────────────────────
    elementos.append(Paragraph('Atenciones Recientes', seccion_e))
    ar_data = [['Fecha', 'Cliente', 'Servicio', 'Técnico', 'Estado']]
    for r in reportes_recientes:
        ar_data.append([
            r.fecha.strftime('%d/%m/%Y'),
            r.cliente.nombre_completo,
            r.servicio.nombre,
            r.tecnico.nombre_completo if r.tecnico else 'Sin Asignar',
            r.estado.nombre if r.estado else '—',
        ])
    ar_tabla = Table(ar_data, colWidths=[2.2*cm, 3.8*cm, 4.5*cm, 3.5*cm, 2.5*cm])
    ar_tabla.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,0), AZUL),
        ('TEXTCOLOR',     (0,0), (-1,0), BLANCO),
        ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0,0), (-1,-1), 8),
        ('ROWBACKGROUNDS',(0,1), (-1,-1), [BLANCO, AZUL_CLARO]),
        ('GRID',          (0,0), (-1,-1), 0.5, GRIS_BORDE),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 5),
        ('WORDWRAP',      (0,0), (-1,-1), True),
    ]))
    elementos.append(ar_tabla)
    elementos.append(Spacer(1, 16))
    elementos.append(HRFlowable(width='100%', thickness=0.5, color=GRIS_BORDE))
    elementos.append(Spacer(1, 4))
    elementos.append(Paragraph(f'ServiTech · Informe generado automáticamente el {hoy.strftime("%d/%m/%Y")} · Confidencial', pie_e))

    doc.build(elementos)
    buffer.seek(0)

    from django.http import FileResponse
    nombre_archivo = f'ServiTech_Reporte_{hoy.strftime("%Y%m%d")}.pdf'
    return FileResponse(buffer, as_attachment=True, filename=nombre_archivo, content_type='application/pdf')


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_exportar_analitico_excel(request):
    """Exporta el informe analítico completo a Excel con openpyxl."""
    if not request.user.es_admin:
        return redirect('home')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    hoy = date.today()
    citas_mes_qs = Cita.objects.filter(fecha__year=hoy.year, fecha__month=hoy.month)
    total       = citas_mes_qs.count()
    finalizadas = citas_mes_qs.filter(estado__nombre__iexact='Finalizada').count()
    canceladas  = citas_mes_qs.filter(estado__nombre__iexact='Cancelada').count()
    tasa_exito  = f"{round(finalizadas / total * 100, 1)}%" if total > 0 else "0%"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe Analítico"
    ws.views.sheetView[0].showGridLines = True

    font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_body = Font(name='Arial', size=10)
    font_bold = Font(name='Arial', size=10, bold=True)
    fill_navy = PatternFill(start_color='002B75', end_color='002B75', fill_type='solid')
    fill_gray = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Banner
    ws.merge_cells('A1:F2')
    title_cell = ws['A1']
    title_cell.value = f"INFORME ANALÍTICO DE RENDIMIENTO — SERVITECH ({hoy.strftime('%B %Y').upper()})"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = align_center

    # 2. Resumen KPIs
    ws['A4'] = "Resumen de Métricas del Mes"
    ws['A4'].font = Font(name='Arial', size=11, bold=True, color='002B75')

    metrics = [
        ("Citas Totales", total),
        ("Servicios Finalizados", finalizadas),
        ("Citas Canceladas", canceladas),
        ("Tasa de Éxito", tasa_exito)
    ]
    for idx, (m_name, m_val) in enumerate(metrics, start=5):
        ws[f'A{idx}'] = m_name
        ws[f'A{idx}'].font = font_bold
        ws[f'A{idx}'].border = border_all
        ws[f'A{idx}'].fill = fill_gray
        
        ws[f'B{idx}'] = m_val
        ws[f'B{idx}'].font = font_body
        ws[f'B{idx}'].border = border_all
        ws[f'B{idx}'].alignment = align_center

    # 3. Listado de Citas Recientes
    ws['A10'] = "Detalle de Atenciones Recientes"
    ws['A10'].font = Font(name='Arial', size=11, bold=True, color='002B75')

    headers = ["#", "Fecha", "Cliente", "Servicio", "Técnico", "Estado"]
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    citas_recientes = Cita.objects.select_related('cliente', 'servicio', 'estado', 'tecnico').order_by('-fecha_creacion')[:20]
    row_num = 12
    for idx, c in enumerate(citas_recientes, start=1):
        ws.cell(row=row_num, column=1, value=f"{idx:02d}").alignment = align_center
        ws.cell(row=row_num, column=2, value=c.fecha.strftime('%d/%m/%Y')).alignment = align_center
        ws.cell(row=row_num, column=3, value=c.cliente.nombre_completo)
        ws.cell(row=row_num, column=4, value=c.servicio.nombre if c.servicio else '—')
        ws.cell(row=row_num, column=5, value=c.tecnico.nombre_completo if c.tecnico else 'Sin Asignar')
        ws.cell(row=row_num, column=6, value=c.estado.nombre if c.estado else '—').alignment = align_center

        for col_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_body
            cell.border = border_all
            if idx % 2 == 0:
                cell.fill = fill_gray
        row_num += 1

    for col_idx in range(1, 7):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Informe_Analitico_{hoy.strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
#  TÉCNICOS (ADMIN)
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_tecnicos(request):
    if not request.user.es_admin:
        return redirect('home')

    from ..models import Cita, EstadoCita, PerfilTecnico
    from django.utils import timezone
    import uuid

    hoy = timezone.now().date()

    # ══════════════════════════════════════════════════════════════════════════
    #  POST: Acciones
    # ══════════════════════════════════════════════════════════════════════════
    if request.method == 'POST':
        accion = request.POST.get('accion', '')

        # ── 1. Crear nuevo técnico + PerfilTecnico ────────────────────────────
        if accion == 'crear_tecnico':
            nombre_completo = request.POST.get('nombre_completo', '').strip()
            correo          = request.POST.get('correo', '').strip()
            telefono        = request.POST.get('telefono', '').strip()
            especialidad    = request.POST.get('especialidad', 'tecnico_general')
            observaciones   = request.POST.get('observaciones', '').strip()
            usuario_id      = request.POST.get('usuario_existente_id', '').strip()

            # Validaciones
            if not nombre_completo:
                messages.error(request, 'El nombre completo es obligatorio.')
                return redirect('admin_tecnicos')
            if not especialidad:
                messages.error(request, 'Selecciona una especialidad.')
                return redirect('admin_tecnicos')

            if usuario_id:
                # Vincular usuario existente
                try:
                    tecnico = Usuario.objects.get(pk=usuario_id)
                    if tecnico.rol != Usuario.Rol.TECNICO:
                        tecnico.rol = Usuario.Rol.TECNICO
                        tecnico.save(update_fields=['rol'])
                    creado = False
                except Usuario.DoesNotExist:
                    messages.error(request, 'El usuario seleccionado no existe.')
                    return redirect('admin_tecnicos')
            else:
                # Crear nuevo usuario
                if not correo:
                    messages.error(request, 'El correo es obligatorio para crear un nuevo técnico.')
                    return redirect('admin_tecnicos')
                if Usuario.objects.filter(correo__iexact=correo).exists():
                    messages.error(request, f'Ya existe un usuario con el correo {correo}.')
                    return redirect('admin_tecnicos')
                password_nuevo = request.POST.get('password', '').strip()
                if not password_nuevo:
                    password_nuevo = uuid.uuid4().hex[:10]
                    
                tecnico = Usuario.objects.create_user(
                    correo=correo,
                    password=password_nuevo,
                    nombre_completo=nombre_completo,
                    telefono=telefono or None,
                    rol=Usuario.Rol.TECNICO,
                    activo=True,
                )
                
                genero = request.POST.get('genero')
                if genero == 'mujer':
                    tecnico.genero = 'F'
                elif genero == 'hombre':
                    tecnico.genero = 'M'
                tecnico.save(update_fields=['genero'])
                
                creado = True

            # Crear o actualizar el PerfilTecnico
            perfil, _ = PerfilTecnico.objects.update_or_create(
                tecnico=tecnico,
                defaults={
                    'especialidad':   especialidad,
                    'observaciones':  observaciones or None,
                }
            )

            # Guardar horarios seleccionados
            from turnos.models.horarios import HorarioTecnico
            from datetime import datetime
            dias_laborales = request.POST.getlist('dias_laborales')
            hora_inicio_str = request.POST.get('hora_inicio', '08:00')
            hora_fin_str = request.POST.get('hora_fin', '18:00')
            
            if not dias_laborales:
                dias_laborales = ['0', '1', '2', '3', '4', '5'] # Lunes a Sábado por defecto
                
            HorarioTecnico.objects.filter(tecnico=tecnico).delete()
            
            try:
                hi = datetime.strptime(hora_inicio_str, '%H:%M').time()
                hf = datetime.strptime(hora_fin_str, '%H:%M').time()
                for dia_str in dias_laborales:
                    HorarioTecnico.objects.create(
                        tecnico=tecnico,
                        dia_semana=int(dia_str),
                        hora_inicio=hi,
                        hora_fin=hf
                    )
            except Exception as e:
                pass

            msg = f'Técnico {tecnico.nombre_completo} {"creado" if creado else "actualizado"} exitosamente.'
            if creado:
                msg += f' Contraseña temporal: {password_nuevo} (cámbiela al primer inicio de sesión).'
            messages.success(request, msg)

        # ── 2. Editar perfil de un técnico existente ──────────────────────────
        elif accion == 'editar_perfil':
            tecnico_id     = request.POST.get('tecnico_id')
            especialidad   = request.POST.get('especialidad', 'tecnico_general')
            observaciones  = request.POST.get('observaciones', '').strip()
            genero         = request.POST.get('genero')

            if not especialidad:
                messages.error(request, 'Selecciona una especialidad.')
                return redirect('admin_tecnicos')
            try:
                tecnico = Usuario.objects.get(pk=tecnico_id, rol=Usuario.Rol.TECNICO)
                if genero == 'hombre':
                    tecnico.genero = 'M'
                    tecnico.save(update_fields=['genero'])
                elif genero == 'mujer':
                    tecnico.genero = 'F'
                    tecnico.save(update_fields=['genero'])
                    
                perfil, _ = PerfilTecnico.objects.update_or_create(
                    tecnico=tecnico,
                    defaults={
                        'especialidad':   especialidad,
                        'observaciones':  observaciones or None,
                    }
                )

                # Guardar horarios seleccionados
                from turnos.models.horarios import HorarioTecnico
                from datetime import datetime
                dias_laborales = request.POST.getlist('dias_laborales')
                hora_inicio_str = request.POST.get('hora_inicio', '08:00')
                hora_fin_str = request.POST.get('hora_fin', '18:00')
                
                HorarioTecnico.objects.filter(tecnico=tecnico).delete()
                
                try:
                    hi = datetime.strptime(hora_inicio_str, '%H:%M').time()
                    hf = datetime.strptime(hora_fin_str, '%H:%M').time()
                    for dia_str in dias_laborales:
                        HorarioTecnico.objects.create(
                            tecnico=tecnico,
                            dia_semana=int(dia_str),
                            hora_inicio=hi,
                            hora_fin=hf
                        )
                except Exception as e:
                    pass
                messages.success(request, f'Perfil de {tecnico.nombre_completo} actualizado exitosamente.')
            except Usuario.DoesNotExist:
                messages.error(request, 'Técnico no encontrado.')

        # ── 3. Asignar técnico a una cita ─────────────────────────────────────
        elif accion == 'asignar_tecnico':
            tecnico_id = request.POST.get('tecnico_id')
            cita_id    = request.POST.get('cita_id')
            if tecnico_id and cita_id:
                try:
                    cita    = Cita.objects.get(pk=cita_id)
                    tecnico = Usuario.objects.get(pk=tecnico_id, rol=Usuario.Rol.TECNICO)
                    cita.tecnico = tecnico
                    cita.save(update_fields=['tecnico'])
                    messages.success(request, f'Técnico {tecnico.nombre_completo} asignado a la cita #{cita_id}.')
                except (Cita.DoesNotExist, Usuario.DoesNotExist):
                    messages.error(request, 'No se pudo completar la asignación.')

        return redirect('admin_tecnicos')

    # ══════════════════════════════════════════════════════════════════════════
    #  GET: Preparar datos
    # ══════════════════════════════════════════════════════════════════════════
    # Búsqueda global desde topbar
    search_query = request.GET.get('search', '').strip()

    tecnicos_qs = Usuario.objects.filter(
        rol=Usuario.Rol.TECNICO
    ).select_related('perfil_tecnico').order_by('nombre_completo')

    if search_query:
        from django.db.models import Q
        tecnicos_qs = tecnicos_qs.filter(
            Q(nombre_completo__icontains=search_query) |
            Q(correo__icontains=search_query) |
            Q(telefono__icontains=search_query)
        )

    tecnicos = tecnicos_qs

    # ── Enriquecer cada técnico con datos de monitoreo ───────────────────────
    tecnicos_monitoreo = []
    for t in tecnicos:
        citas_hoy = Cita.objects.filter(
            tecnico=t, fecha=hoy
        ).select_related('estado', 'cliente', 'servicio').order_by('hora_inicio')

        cita_activa = citas_hoy.exclude(
            estado__nombre__in=['CANCELADA', 'FINALIZADA']
        ).first()

        # Obtener perfil si existe
        try:
            perfil = t.perfil_tecnico
        except Exception:
            perfil = None

        if not t.activo:
            estado_monitor = 'inactivo'
        elif cita_activa:
            nombre_estado = (cita_activa.estado.nombre if cita_activa.estado else '').lower()
            if 'proceso' in nombre_estado or 'atendi' in nombre_estado or 'en curso' in nombre_estado:
                estado_monitor = 'en_proceso'
            elif 'pausa' in nombre_estado or 'novedad' in nombre_estado or 'retraso' in nombre_estado:
                estado_monitor = 'pausado'
            else:
                estado_monitor = 'en_proceso'
        elif perfil and getattr(perfil, 'en_pausa_manual', False):
            estado_monitor = 'pausado'
        else:
            estado_monitor = 'disponible'

        total_hoy       = citas_hoy.count()
        completadas_hoy = citas_hoy.filter(estado__nombre__in=['FINALIZADA']).count()

        tecnicos_monitoreo.append({
            'tecnico':         t,
            'perfil':          perfil,
            'cita_activa':     cita_activa,
            'citas_hoy':       list(citas_hoy[:3]),
            'total_citas_hoy': total_hoy,
            'completadas_hoy': completadas_hoy,
            'estado_monitor':  estado_monitor,
        })

    orden_estado = {'en_proceso': 0, 'disponible': 1, 'pausado': 2, 'inactivo': 3}
    tecnicos_monitoreo.sort(key=lambda x: orden_estado.get(x['estado_monitor'], 4))

    total_tecnicos  = tecnicos.count()
    en_proceso_hoy  = sum(1 for t in tecnicos_monitoreo if t['estado_monitor'] == 'en_proceso')
    disponibles_hoy = sum(1 for t in tecnicos_monitoreo if t['estado_monitor'] == 'disponible')
    pausados_hoy    = sum(1 for t in tecnicos_monitoreo if t['estado_monitor'] == 'pausado')

    citas_sin_tecnico = Cita.objects.filter(
        tecnico__isnull=True, fecha=hoy
    ).exclude(
        estado__nombre__in=['CANCELADA', 'FINALIZADA']
    ).select_related('cliente', 'servicio', 'estado').order_by('hora_inicio')[:5]

    # Usuarios NO-técnicos para el selector de vincular usuario existente
    usuarios_vinculables = Usuario.objects.filter(
        activo=True
    ).exclude(rol=Usuario.Rol.TECNICO).order_by('nombre_completo')[:30]

    context = {
        'tecnicos':            tecnicos,
        'tecnicos_monitoreo':  tecnicos_monitoreo,
        'total_tecnicos':      total_tecnicos,
        'en_proceso_hoy':      en_proceso_hoy,
        'disponibles_hoy':     disponibles_hoy,
        'pausados_hoy':        pausados_hoy,
        'citas_sin_tecnico':   citas_sin_tecnico,
        'hoy':                 hoy,
        'usuarios_vinculables': usuarios_vinculables,
        'niveles':             PerfilTecnico.Nivel.choices,
        'dispositivos':        PerfilTecnico.EspecialidadTecnico.choices,
        'search_query':        search_query,
    }
    return render(request, 'turnos/administracion/admin_tecnicos.html', context)


# ─────────────────────────────────────────────
#  INVENTARIO (ADMIN)
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_inventario(request):
    if not request.user.es_admin:
        return redirect('home')
    from ..models import Repuesto, Inventario

    # ── REGISTRO DE INGRESO DE STOCK (POST) ──────────────────────────────────
    if request.method == 'POST':
        nombre    = request.POST.get('nombre', '').strip()
        categoria = request.POST.get('categoria', 'OTROS').strip()
        cantidad  = request.POST.get('cantidad', '0').strip()
        proveedor = request.POST.get('proveedor', '').strip()
        notas     = request.POST.get('notas', '').strip()
        imagen    = request.FILES.get('imagen')

        # Validaciones básicas
        errores = []
        if not nombre:
            errores.append('El nombre del artículo es obligatorio.')
        try:
            cantidad_int = int(cantidad)
            if cantidad_int < 1:
                raise ValueError
        except (ValueError, TypeError):
            errores.append('La cantidad debe ser un número entero mayor a 0.')

        if errores:
            for e in errores:
                messages.error(request, e)
            return redirect('admin_inventario')

        # Buscar o crear el Repuesto (artículo)
        repuesto, creado = Repuesto.objects.get_or_create(
            nombre__iexact=nombre,
            defaults={
                'nombre':    nombre,
                'categoria': categoria,
                'precio':    0,
                'proveedor': proveedor or None,
                'imagen':    imagen,
                'activo':    True,
            }
        )

        if not creado:
            # Si ya existía, actualizar proveedor/categoría/imagen si se enviaron
            actualizado = False
            if proveedor and repuesto.proveedor != proveedor:
                repuesto.proveedor = proveedor
                actualizado = True
            if categoria and repuesto.categoria != categoria:
                repuesto.categoria = categoria
                actualizado = True
            if imagen:
                repuesto.imagen = imagen
                actualizado = True
            if actualizado:
                repuesto.save(update_fields=['proveedor', 'categoria', 'imagen'])

        # Actualizar stock en el artículo
        repuesto.stock += cantidad_int
        repuesto.save(update_fields=['stock'])

        # Registrar movimiento en tabla Inventario
        Inventario.objects.create(
            repuesto=repuesto,
            cantidad=cantidad_int,
            tipo='ENTRADA',
            usuario=request.user,
            motivo=notas or f'Ingreso manual - Proveedor: {proveedor}' if proveedor else 'Ingreso manual',
        )

        accion = 'creado y registrado' if creado else 'actualizado'
        messages.success(
            request,
            f'✅ ¡{nombre}! Stock {accion}: +{cantidad_int} unidades. Total actual: {repuesto.stock} u.'
        )
        return redirect('admin_inventario')

    # ── GET: preparar datos para el template ─────────────────────────────────
    repuestos_base = Repuesto.objects.filter(activo=True).order_by('categoria', 'nombre')

    # Búsqueda global desde topbar
    search_query = request.GET.get('search', '').strip()
    if search_query:
        from django.db.models import Q
        repuestos = repuestos_base.filter(
            Q(nombre__icontains=search_query) |
            Q(categoria__icontains=search_query) |
            Q(proveedor__icontains=search_query)
        )
    else:
        repuestos = repuestos_base

    # Métricas (sobre base completa, no sobre búsqueda)
    stock_critico = repuestos_base.filter(stock__lt=5).count()
    stock_agotado = repuestos_base.filter(stock=0).count()
    total_items   = repuestos_base.count()
    valorizacion  = sum(r.stock * r.precio for r in repuestos_base if r.stock > 0)

    # Alertas de stock (< 5 unidades)
    alertas_stock = repuestos_base.filter(stock__lt=5).order_by('stock')[:6]

    # Últimos 10 ingresos registrados
    from django.utils import timezone
    from datetime import timedelta
    ultimos_ingresos = Inventario.objects.filter(
        tipo='ENTRADA'
    ).select_related('repuesto', 'usuario').order_by('-fecha')[:10]

    # Ingresos de hoy
    hoy = timezone.now().date()
    ingresos_hoy   = Inventario.objects.filter(tipo='ENTRADA', fecha__date=hoy).count()
    proveedores_hoy = Repuesto.objects.filter(
        inventarios__tipo='ENTRADA',
        inventarios__fecha__date=hoy
    ).values('proveedor').distinct().count()

    context = {
        'repuestos':        repuestos,
        'stock_critico':    stock_critico,
        'stock_agotado':    stock_agotado,
        'total_items':      total_items,
        'valorizacion':     valorizacion,
        'alertas_stock':    alertas_stock,
        'ultimos_ingresos': ultimos_ingresos,
        'ingresos_hoy':     ingresos_hoy,
        'proveedores_hoy':  proveedores_hoy,
        'categorias':       Repuesto.CATEGORIA_CHOICES,
        'search_query':     search_query,
    }
    return render(request, 'turnos/administracion/admin_inventario.html', context)


# ─────────────────────────────────────────────
#  HISTORIAL COMPLETO DE INVENTARIO
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_historial_inventario(request):
    """Historial completo de movimientos de inventario con filtros."""
    if not request.user.es_admin:
        return redirect('home')

    from ..models import Repuesto, Inventario
    from django.utils import timezone
    from django.core.paginator import Paginator

    # ── Filtros desde GET ──────────────────────────────────────────────────────
    tipo_filtro      = request.GET.get('tipo', '')       # ENTRADA | SALIDA | AJUSTE
    categoria_filtro = request.GET.get('categoria', '')  # categoria de repuesto
    busqueda         = request.GET.get('q', '').strip()
    fecha_desde      = request.GET.get('fecha_desde', '')
    fecha_hasta      = request.GET.get('fecha_hasta', '')

    movimientos = Inventario.objects.select_related(
        'repuesto', 'usuario'
    ).order_by('-fecha')

    if tipo_filtro:
        movimientos = movimientos.filter(tipo=tipo_filtro)
    if categoria_filtro:
        movimientos = movimientos.filter(repuesto__categoria=categoria_filtro)
    if busqueda:
        movimientos = movimientos.filter(repuesto__nombre__icontains=busqueda)
    if fecha_desde:
        movimientos = movimientos.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        movimientos = movimientos.filter(fecha__date__lte=fecha_hasta)

    # ── Paginación (20 por página) ─────────────────────────────────────────────
    paginator   = Paginator(movimientos, 20)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    # ── Resumen global (sin filtros para tener totales reales) ─────────────────
    total_entradas = Inventario.objects.filter(tipo='ENTRADA').count()
    total_salidas  = Inventario.objects.filter(tipo='SALIDA').count()
    total_ajustes  = Inventario.objects.filter(tipo='AJUSTE').count()
    total_registros = Inventario.objects.count()

    # Artículos únicos con movimientos
    articulos_con_movimiento = Inventario.objects.values('repuesto').distinct().count()

    context = {
        'page_obj':       page_obj,
        'movimientos':    page_obj.object_list,
        'total_entradas': total_entradas,
        'total_salidas':  total_salidas,
        'total_ajustes':  total_ajustes,
        'total_registros':total_registros,
        'articulos_con_movimiento': articulos_con_movimiento,
        # Filtros activos para mantenerlos en el template
        'tipo_filtro':      tipo_filtro,
        'categoria_filtro': categoria_filtro,
        'busqueda':         busqueda,
        'fecha_desde':      fecha_desde,
        'fecha_hasta':      fecha_hasta,
        'categorias':       Repuesto.CATEGORIA_CHOICES,
        'tipos':            Inventario.TIPO_MOVIMIENTO_CHOICES,
    }
    return render(request, 'turnos/administracion/admin_historial_inventario.html', context)


@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_exportar_usuarios_excel(request):
    """Exporta el directorio de usuarios a Excel con formato institucional."""
    if not request.user.es_admin:
        return redirect('home')
    
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from turnos.models import Usuario

    qs = Usuario.objects.all().order_by('-fecha_registro')

    # Filtros desde GET
    search = request.GET.get('search', '').strip()
    rol = request.GET.get('rol', '').strip()
    estado_param = request.GET.get('estado', '').strip()

    if search:
        from django.db.models import Q
        qs = qs.filter(Q(nombre_completo__icontains=search) | Q(correo__icontains=search) | Q(telefono__icontains=search))
    if rol and rol != 'TODOS':
        qs = qs.filter(rol=rol)
    if estado_param:
        if estado_param == 'ACTIVO':
            qs = qs.filter(activo=True)
        elif estado_param == 'INACTIVO':
            qs = qs.filter(activo=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directorio Usuarios"

    # Mostrar cuadrícula
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_title = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    font_body = Font(name='Arial', size=10)
    font_bold = Font(name='Arial', size=10, bold=True)
    
    fill_navy = PatternFill(start_color='002B75', end_color='002B75', fill_type='solid')
    fill_gray = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    thin_side = Side(border_style="thin", color="CBD5E1")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # 1. Banner Superior
    ws.merge_cells('A1:F2')
    title_cell = ws['A1']
    title_cell.value = "DIRECTORIO DE USUARIOS REGISTRADOS - SERVITECH"
    title_cell.font = font_title
    title_cell.fill = fill_navy
    title_cell.alignment = align_center

    # 2. Encabezados
    headers = ['ID', 'Nombre Completo', 'Correo Electrónico', 'Teléfono', 'Rol', 'Estado']
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_all

    # 3. Filas de datos
    row_num = 5
    for idx, u in enumerate(qs, start=1):
        ws.cell(row=row_num, column=1, value=u.pk).alignment = align_center
        ws.cell(row=row_num, column=2, value=u.nombre_completo).alignment = align_left
        ws.cell(row=row_num, column=3, value=u.correo).alignment = align_left
        ws.cell(row=row_num, column=4, value=u.telefono or '—').alignment = align_center
        ws.cell(row=row_num, column=5, value=u.get_rol_display()).alignment = align_center
        
        est_txt = 'Activo' if u.activo else 'Inactivo'
        ws.cell(row=row_num, column=6, value=est_txt).alignment = align_center

        for col_idx in range(1, 7):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = font_body
            cell.border = border_all
            if idx % 2 == 0:
                cell.fill = fill_gray
        row_num += 1

    # Autoajustar ancho de columnas
    for col_idx in range(1, 7):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="directorio_usuarios.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────────────────────────
#  MI PERFIL (ADMINISTRADOR)
# ─────────────────────────────────────────────
@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def admin_perfil(request):
    """Perfil del Administrador: actualización de datos, foto y contraseña."""
    if not request.user.es_admin:
        return redirect('home')

    if request.method == 'POST':
        form = EditarPerfilForm(
            request.POST,
            request.FILES,
            instance=request.user,
            current_user=request.user
        )
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.save()
            # Si se cambió la contraseña re-autenticamos la sesión para no cerrar la sesión
            from django.contrib.auth import update_session_auth_hash
            if form.cleaned_data.get('password_new'):
                update_session_auth_hash(request, usuario)
            messages.success(request, '¡Tu perfil fue actualizado con éxito!')
            return redirect('admin_perfil')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{error}")
    else:
        form = EditarPerfilForm(instance=request.user, current_user=request.user)

    return render(request, 'turnos/administracion/admin_perfil.html', {
        'form': form,
        'usuario': request.user,
    })


# ─────────────────────────────────────────────
#  TOGGLE DE PAUSA MANUAL (TÉCNICO)
# ─────────────────────────────────────────────
from django.http import JsonResponse

@login_required
@rol_requerido([Usuario.Rol.ADMINISTRADOR])
def tecnico_toggle_pausa(request):
    """
    Alterna el estado de pausa manual del técnico. Soporta AJAX.
    """
    if request.method == 'POST' and request.user.rol == Usuario.Rol.TECNICO:
        try:
            perfil = request.user.perfil_tecnico
            
            if not perfil.en_pausa_manual:
                from django.utils import timezone
                hoy = timezone.now().date()
                citas_activas = request.user.citas_tecnico.filter(
                    fecha=hoy
                ).exclude(estado__nombre__in=['CANCELADA', 'FINALIZADA'])
                
                if citas_activas.exists():
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
                        return JsonResponse({'success': False, 'error': "No puedes pausar tu turno mientras tengas citas activas o pendientes."})
                    messages.error(request, "No puedes pausar tu turno mientras tengas citas activas o pendientes.")
                    return redirect('dashboard_tecnico')
                    
            perfil.en_pausa_manual = not perfil.en_pausa_manual
            perfil.save(update_fields=['en_pausa_manual'])
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
                return JsonResponse({'success': True, 'pausado': perfil.en_pausa_manual})
                
            estado_msg = "EN PAUSA" if perfil.en_pausa_manual else "DISPONIBLE"
            messages.success(request, f"Turno actualizado. Ahora estás {estado_msg}.")
        except Exception as e:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.content_type == 'application/json':
                return JsonResponse({'success': False, 'error': str(e)})
            messages.error(request, f"Error al cambiar estado: {e}")
            
    return redirect('dashboard_tecnico')

def api_estado_tecnicos(request):
    """
    Devuelve el estado de todos los técnicos para actualizar el panel de admin en tiempo real.
    """
    tecnicos = Usuario.objects.filter(rol=Usuario.Rol.TECNICO, is_active=True)
    data = []
    
    from django.utils import timezone
    hoy = timezone.localdate()
    ahora = timezone.localtime().time()
    
    for t in tecnicos:
        perfil = t.perfil_tecnico if hasattr(t, 'perfil_tecnico') else None
        estado = 'disponible'
        
        if perfil and perfil.en_pausa_manual:
            estado = 'pausado'
        else:
            cita_activa = t.citas_tecnico.filter(
                fecha=hoy, 
                hora_inicio__lte=ahora, 
                hora_fin__gte=ahora
            ).exclude(estado__nombre__in=['CANCELADA', 'FINALIZADA']).first()
            if cita_activa:
                estado = 'en_proceso'
                
        data.append({
            'id': t.pk,
            'estado': estado
        })
        
    return JsonResponse({'success': True, 'tecnicos': data})


@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_historial(request):
    hoy = date.today()

    # ── KPIs ──
    qs_total = Cita.objects.filter(tecnico=request.user)
    
    citas_hoy      = qs_total.filter(fecha=hoy).count()
    pendientes     = qs_total.filter(estado__nombre__in=['PENDIENTE', 'CONFIRMADA', 'RETRASADA']).count()
    en_proceso     = qs_total.filter(estado__nombre__in=['EN_REPARACION']).count()
    completadas    = qs_total.filter(estado__nombre__in=['FINALIZADA']).count()

    # ── Filtros desde GET ──
    qs = qs_total.select_related('cliente', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    filtro_cliente  = request.GET.get('cliente', '').strip()
    filtro_estado   = request.GET.get('estado', '').strip()
    filtro_fecha_ini = request.GET.get('fecha_ini', '').strip()
    filtro_fecha_fin = request.GET.get('fecha_fin', '').strip()

    if filtro_cliente:
        qs = qs.filter(cliente__nombre_completo__icontains=filtro_cliente)
    if filtro_estado:
        qs = qs.filter(estado__nombre__iexact=filtro_estado)
    if filtro_fecha_ini:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__gte=datetime.strptime(filtro_fecha_ini, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filtro_fecha_fin:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__lte=datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Paginación simple (25 por página)
    from django.core.paginator import Paginator
    paginator   = Paginator(qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    estados = EstadoCita.objects.all().order_by('nombre')

    context = {
        'citas':           page_obj,
        'page_obj':        page_obj,
        'total_citas':     qs.count(),
        'citas_hoy':       citas_hoy,
        'pendientes':      pendientes,
        'en_proceso':      en_proceso,
        'completadas':     completadas,
        'estados':         estados,
        'filtro_cliente':  filtro_cliente,
        'filtro_estado':   filtro_estado,
        'filtro_fecha_ini': filtro_fecha_ini,
        'filtro_fecha_fin': filtro_fecha_fin,
    }
    return render(request, 'turnos/tecnico/tecnico_historial.html', context)

@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def exportar_historial_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime

    qs = Cita.objects.filter(tecnico=request.user).select_related('cliente', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    filtro_cliente  = request.GET.get('cliente', '').strip()
    filtro_estado   = request.GET.get('estado', '').strip()
    filtro_fecha_ini = request.GET.get('fecha_ini', '').strip()
    filtro_fecha_fin = request.GET.get('fecha_fin', '').strip()

    if filtro_cliente:
        qs = qs.filter(cliente__nombre_completo__icontains=filtro_cliente)
    if filtro_estado:
        qs = qs.filter(estado__nombre__iexact=filtro_estado)
    if filtro_fecha_ini:
        try:
            qs = qs.filter(fecha__gte=datetime.strptime(filtro_fecha_ini, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filtro_fecha_fin:
        try:
            qs = qs.filter(fecha__lte=datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Citas"

    # Estilos
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003399', end_color='003399', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin', color='CBD5E1'),
        right=Side(border_style='thin', color='CBD5E1'),
        top=Side(border_style='thin', color='CBD5E1'),
        bottom=Side(border_style='thin', color='CBD5E1')
    )

    headers = ['ID', 'Cliente', 'Servicio', 'Fecha', 'Hora', 'Estado', 'Dispositivo']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['G'].width = 25

    row_num = 2
    for cita in qs:
        fecha_str = cita.fecha.strftime('%d/%m/%Y') if cita.fecha else ''
        hora_str = cita.hora_inicio.strftime('%H:%M') if cita.hora_inicio else ''
        ws.append([
            cita.pk,
            cita.cliente.nombre_completo if cita.cliente else 'N/A',
            cita.servicio.nombre if cita.servicio else 'N/A',
            fecha_str,
            hora_str,
            cita.estado.nombre if cita.estado else 'N/A',
            cita.servicio.tipo_dispositivo if cita.servicio else 'N/A'
        ])
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
        row_num += 1

    timestamp = timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="historial_tecnico_{timestamp}.xlsx"'
    wb.save(response)
    return response



@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_cliente_historial(request, cliente_id):
    """Devuelve el historial clínico completo de los dispositivos de un cliente en formato JSON."""
    try:
        cliente = Usuario.objects.get(id_usuario=cliente_id, rol=Usuario.Rol.CLIENTE)
    except Usuario.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Cliente no encontrado'})

    from turnos.models.dispositivos import Dispositivo
    from turnos.models.citas import Cita

    # Obtener dispositivos del cliente
    dispositivos = Dispositivo.objects.filter(cliente=cliente).order_by('-fecha_registro')
    
    # También buscamos citas que no tengan dispositivo asignado y las agrupamos bajo "Dispositivo Genérico"
    citas_sin_dispositivo = Cita.objects.filter(cliente=cliente, dispositivo__isnull=True).order_by('-fecha', '-hora_inicio')

    data_dispositivos = []

    for disp in dispositivos:
        citas_disp = Cita.objects.filter(dispositivo=disp).order_by('-fecha', '-hora_inicio')
        if citas_disp.exists():
            citas_data = [{
                'id': c.id,
                'servicio': c.servicio.nombre,
                'fecha': c.fecha.strftime('%d/%m/%Y'),
                'estado': c.estado.nombre.upper() if c.estado else 'PENDIENTE',
                'observaciones': c.observaciones or ''
            } for c in citas_disp]
            
            data_dispositivos.append({
                'id': disp.id,
                'marca': disp.marca,
                'modelo': disp.modelo,
                'imei_serial': disp.imei_serial,
                'citas': citas_data
            })

    if citas_sin_dispositivo.exists():
        citas_data = [{
            'id': c.id,
            'servicio': c.servicio.nombre,
            'fecha': c.fecha.strftime('%d/%m/%Y'),
            'estado': c.estado.nombre.upper() if c.estado else 'PENDIENTE',
            'observaciones': c.observaciones or ''
        } for c in citas_sin_dispositivo]
        
        data_dispositivos.append({
            'id': 0,
            'marca': 'Equipo',
            'modelo': 'Generico',
            'imei_serial': 'No registrado',
            'citas': citas_data
        })

    return JsonResponse({'success': True, 'dispositivos': data_dispositivos})


@login_required
@rol_requerido([Usuario.Rol.TECNICO])
def tecnico_clientes_historial_general(request):
    """Devuelve el historial general de reparaciones (todas las citas finalizadas) del técnico."""
    from turnos.models.citas import Cita
    # Solo citas finalizadas del técnico
    citas = Cita.objects.filter(
        tecnico=request.user,
        estado__nombre__icontains='FINALIZ'
    ).select_related('cliente', 'servicio', 'dispositivo').order_by('-fecha', '-hora_inicio')[:50] # Top 50 para no saturar

    data = []
    for c in citas:
        data.append({
            'id': c.id,
            'cliente': c.cliente.nombre_completo,
            'fecha': c.fecha.strftime('%d/%m/%Y'),
            'servicio': c.servicio.nombre,
            'dispositivo': f"{c.dispositivo.marca} {c.dispositivo.modelo}" if c.dispositivo else "Equipo Genérico"
        })
        
    return JsonResponse({'success': True, 'historial': data})

@login_required
def admin_toggle_pausa_tecnico(request, tecnico_id):
    """
    Alterna el estado de pausa de un técnico por parte del administrador.
    """
    from django.shortcuts import get_object_or_404
    from turnos.models.usuarios import Usuario
    
    if request.user.rol != Usuario.Rol.ADMINISTRADOR:
        return JsonResponse({'success': False, 'error': 'No autorizado'}, status=403)
        
    try:
        tecnico = get_object_or_404(Usuario, pk=tecnico_id)
        if not hasattr(tecnico, 'perfil_tecnico'):
            return JsonResponse({'success': False, 'error': 'El usuario no tiene perfil de técnico'})
            
        perfil = tecnico.perfil_tecnico
        is_pausing = not perfil.en_pausa_manual
        perfil.en_pausa_manual = is_pausing
        perfil.save(update_fields=['en_pausa_manual'])
        
        if is_pausing:
            from django.utils import timezone
            from turnos.models.notificaciones import Notificacion
            from turnos.models.citas import Cita
            
            today = timezone.now().date()
            citas_afectadas = Cita.objects.filter(
                tecnico=tecnico,
                fecha__gte=today,
                estado__nombre__in=['PENDIENTE', 'CONFIRMADA', 'RETRASADA']
            )
            for cita in citas_afectadas:
                especialidad_disp = perfil.get_especialidad_display()
                mensaje = (
                    f"Tu cita programada para el {cita.fecha.strftime('%d/%m/%Y')} "
                    f"a las {cita.hora_inicio.strftime('%H:%M')} no podrá ser atendida "
                    f"debido a una suspensión temporal del técnico {tecnico.nombre_completo}. "
                    f"Puedes reagendar tu cita con otro técnico de la misma especialidad ({especialidad_disp})."
                )
                Notificacion.objects.create(
                    usuario=cita.cliente,
                    cita=cita,
                    tipo='SUSPENSION_TECNICO',
                    mensaje=mensaje
                )
        
        return JsonResponse({'success': True, 'pausado': perfil.en_pausa_manual})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ─────────────────────────────────────────────
#  API: BÚSQUEDA GLOBAL (ADMIN)
# ─────────────────────────────────────────────
@login_required
def api_busqueda_global(request):
    """Devuelve conteos de resultados por módulo para el buscador global del topbar admin."""
    if not request.user.es_admin:
        return JsonResponse({'error': 'No autorizado'}, status=403)

    q = request.GET.get('q', '').strip()
    if not q or len(q) < 2:
        return JsonResponse({'resultados': {}})

    from django.db.models import Q
    from ..models import Repuesto, Inventario

    resultados = {}

    # Usuarios (clientes)
    usuarios_count = Usuario.objects.filter(
        Q(nombre_completo__icontains=q) | Q(correo__icontains=q) | Q(telefono__icontains=q),
        rol=Usuario.Rol.CLIENTE
    ).count()
    resultados['usuarios'] = {
        'count': usuarios_count,
        'url': f"/admin-panel/usuarios/?search={q}",
        'label': 'Usuarios',
        'icon': '👤',
    }

    # Técnicos
    tecnicos_count = Usuario.objects.filter(
        Q(nombre_completo__icontains=q) | Q(correo__icontains=q) | Q(telefono__icontains=q),
        rol=Usuario.Rol.TECNICO
    ).count()
    resultados['tecnicos'] = {
        'count': tecnicos_count,
        'url': f"/admin-panel/tecnicos/?search={q}",
        'label': 'Técnicos',
        'icon': '🔧',
    }

    # Citas
    citas_count = Cita.objects.filter(
        Q(cliente__nombre_completo__icontains=q) |
        Q(tecnico__nombre_completo__icontains=q) |
        Q(servicio__nombre__icontains=q)
    ).count()
    resultados['citas'] = {
        'count': citas_count,
        'url': f"/admin-panel/citas/?cliente={q}",
        'label': 'Citas',
        'icon': '📅',
    }

    # Servicios / Catálogo
    servicios_count = Servicio.objects.filter(
        Q(nombre__icontains=q) | Q(descripcion__icontains=q)
    ).count()
    resultados['servicios'] = {
        'count': servicios_count,
        'url': f"/admin-panel/servicios/?search={q}",
        'label': 'Catálogo',
        'icon': '📋',
    }

    # Inventario / Repuestos
    inventario_count = Repuesto.objects.filter(
        Q(nombre__icontains=q) | Q(categoria__icontains=q) | Q(proveedor__icontains=q)
    ).count()
    resultados['inventario'] = {
        'count': inventario_count,
        'url': f"/admin-panel/inventario/?search={q}",
        'label': 'Inventario',
        'icon': '📦',
    }

    return JsonResponse({'resultados': resultados, 'q': q})

