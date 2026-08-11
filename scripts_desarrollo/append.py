import codecs

code = """
@login_required
def tecnico_historial(request):
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    hoy = date.today()

    # ── KPIs ──
    qs_total = Cita.objects.filter(tecnico=request.user)
    
    citas_hoy      = qs_total.filter(fecha=hoy).count()
    pendientes     = qs_total.filter(estado__nombre__iexact='Confirmada').count()
    en_proceso     = qs_total.filter(estado__nombre__iexact='Diagnóstico').count()
    completadas    = qs_total.filter(estado__nombre__iexact='Finalizada').count()

    # ── Filtros desde GET ──
    qs = qs_total.select_related('cliente', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    filtro_cliente  = request.GET.get('cliente', '').strip()
    filtro_estado   = request.GET.get('estado', '').strip()
    filtro_fecha_ini = request.GET.get('fecha_ini', '').strip()
    filtro_fecha_fin = request.GET.get('fecha_fin', '').strip()

    if filtro_cliente:
        qs = qs.filter(cliente__nombre_completo__icontains=filtro_cliente)
    if filtro_estado:
        qs = qs.filter(estado__nombre__iexact=filtro_estado)
    if filtro_fecha_ini:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__gte=datetime.strptime(filtro_fecha_ini, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filtro_fecha_fin:
        try:
            from datetime import datetime
            qs = qs.filter(fecha__lte=datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    # Paginación simple (25 por página)
    from django.core.paginator import Paginator
    paginator   = Paginator(qs, 25)
    page_number = request.GET.get('page', 1)
    page_obj    = paginator.get_page(page_number)

    estados = EstadoCita.objects.all().order_by('nombre')

    context = {
        'citas':           page_obj,
        'page_obj':        page_obj,
        'total_citas':     qs.count(),
        'citas_hoy':       citas_hoy,
        'pendientes':      pendientes,
        'en_proceso':      en_proceso,
        'completadas':     completadas,
        'estados':         estados,
        'filtro_cliente':  filtro_cliente,
        'filtro_estado':   filtro_estado,
        'filtro_fecha_ini': filtro_fecha_ini,
        'filtro_fecha_fin': filtro_fecha_fin,
    }
    return render(request, 'turnos/tecnico/tecnico_historial.html', context)

@login_required
def exportar_historial_excel(request):
    if request.user.rol != Usuario.Rol.TECNICO:
        return redirect('home')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    from django.utils import timezone
    from datetime import datetime

    qs = Cita.objects.filter(tecnico=request.user).select_related('cliente', 'servicio', 'estado').order_by('-fecha', '-hora_inicio')

    filtro_cliente  = request.GET.get('cliente', '').strip()
    filtro_estado   = request.GET.get('estado', '').strip()
    filtro_fecha_ini = request.GET.get('fecha_ini', '').strip()
    filtro_fecha_fin = request.GET.get('fecha_fin', '').strip()

    if filtro_cliente:
        qs = qs.filter(cliente__nombre_completo__icontains=filtro_cliente)
    if filtro_estado:
        qs = qs.filter(estado__nombre__iexact=filtro_estado)
    if filtro_fecha_ini:
        try:
            qs = qs.filter(fecha__gte=datetime.strptime(filtro_fecha_ini, '%Y-%m-%d').date())
        except ValueError:
            pass
    if filtro_fecha_fin:
        try:
            qs = qs.filter(fecha__lte=datetime.strptime(filtro_fecha_fin, '%Y-%m-%d').date())
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Citas"

    # Estilos
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='003399', end_color='003399', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin', color='CBD5E1'),
        right=Side(border_style='thin', color='CBD5E1'),
        top=Side(border_style='thin', color='CBD5E1'),
        bottom=Side(border_style='thin', color='CBD5E1')
    )

    headers = ['ID', 'Cliente', 'Servicio', 'Fecha', 'Hora', 'Estado', 'Dispositivo']
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['G'].width = 25

    row_num = 2
    for cita in qs:
        fecha_str = cita.fecha.strftime('%d/%m/%Y') if cita.fecha else ''
        hora_str = cita.hora_inicio.strftime('%H:%M') if cita.hora_inicio else ''
        ws.append([
            cita.pk,
            cita.cliente.nombre_completo if cita.cliente else 'N/A',
            cita.servicio.nombre if cita.servicio else 'N/A',
            fecha_str,
            hora_str,
            cita.estado.nombre if cita.estado else 'N/A',
            cita.vehiculo or 'N/A'
        ])
        for col_num in range(1, 8):
            ws.cell(row=row_num, column=col_num).border = border
        row_num += 1

    timestamp = timezone.now().strftime('%Y%m%d_%H%M')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="historial_tecnico_{timestamp}.xlsx"'
    wb.save(response)
    return response

"""

with codecs.open(r'c:\Servitech\Servitech-app\turnos\views\dashboard_views.py', 'a', 'utf-8') as f:
    f.write('\n' + code + '\n')
